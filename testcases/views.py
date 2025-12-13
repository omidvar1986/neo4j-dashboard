from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, Http404, HttpResponse
from django.views.decorators.http import require_http_methods
from django.urls import reverse
from bson import ObjectId
from bson.errors import InvalidId
import csv
import html
import io
import json
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import TestCase, TestStep, Section, SectionPermission, ChangeHistory, Project, TestRun, TestRunResult, ensure_section_indexes
from .forms import TestCaseForm


def get_project_or_redirect(project_id):
    """Helper function to get project or return None if not found"""
    try:
        return Project.objects.get(id=ObjectId(project_id), is_active=True)
    except (Project.DoesNotExist, Exception):
        return None


def user_can_edit_sections(user):
    """Check if user can edit sections (admin or has permission)"""
    if user.role == 3:  # Admin
        return True
    try:
        permission = SectionPermission.objects.get(user_id=user.id)
        return permission.can_edit
    except SectionPermission.DoesNotExist:
        return False


def user_can_delete_sections(user):
    """Check if user can delete sections (admin or has permission)"""
    if user.role == 3:  # Admin
        return True
    try:
        permission = SectionPermission.objects.get(user_id=user.id)
        return permission.can_delete
    except SectionPermission.DoesNotExist:
        return False


def record_change(entity_type, entity_id, action, user_id, description=''):
    """Record a change in the change history"""
    try:
        ChangeHistory(
            entity_type=entity_type,
            entity_id=str(entity_id),
            action=action,
            user_id=user_id,
            description=description
        ).save()
    except Exception as e:
        # Don't fail the main operation if history recording fails
        print(f"Error recording change history: {str(e)}")


class TestCaseImportService:
    """Batch helper for importing test cases from spreadsheets/XML."""

    def __init__(self, project, user, batch_size=200):
        self.project = project
        self.user = user
        self.batch_size = batch_size
        self.pending_cases = []
        self.pending_history = []
        self.section_cache = {}
        self.existing_case_keys = set()
        self.stats = {
            'rows_processed': 0,
            'cases_created': 0,
            'steps_created': 0,
            'duplicates_skipped': 0,
            'invalid_rows': 0,
            'sections_created': 0,
            'errors': [],
        }
        # Preload sections for quick lookup
        for section in Section.objects(project=self.project):
            self.section_cache[self._section_key(section.parent, section.name)] = section
        # Preload existing cases (section_id + title) to avoid dereferencing missing sections
        existing_cases = TestCase.objects(project=self.project).only('section', 'title').no_dereference()
        for existing in existing_cases:
            section_id = getattr(existing, 'section_id', None)
            if section_id:
                self.existing_case_keys.add((str(section_id), self._normalize(existing.title)))

    @staticmethod
    def _normalize(value):
        return (value or '').strip().lower()

    def _section_key(self, parent, name):
        parent_id = str(parent.id) if parent else 'root'
        return (parent_id, self._normalize(name))

    def _case_key(self, section, title):
        return (str(section.id), self._normalize(title))

    def _create_section(self, parent, name):
        section = Section(
            project=self.project,
            name=name.strip(),
            parent=parent,
            created_by_id=self.user.id
        )
        section.save()
        self.section_cache[self._section_key(parent, name)] = section
        self.stats['sections_created'] += 1
        return section

    def get_or_create_section_by_path(self, section_path):
        names = [s.strip() for s in section_path.split('>') if s and s.strip()]
        parent_section = None
        for section_name in names:
            key = self._section_key(parent_section, section_name)
            section = self.section_cache.get(key)
            if not section:
                query = Section.objects(project=self.project, name=section_name)
                query = query.filter(parent=parent_section) if parent_section else query.filter(parent__exists=False)
                section = query.first()
                if not section:
                    section = self._create_section(parent_section, section_name)
                else:
                    self.section_cache[key] = section
            parent_section = section
        return parent_section

    def queue_test_case(self, test_case):
        if not test_case.section or not test_case.title:
            self.stats['invalid_rows'] += 1
            return False
        case_key = self._case_key(test_case.section, test_case.title)
        if case_key in self.existing_case_keys:
            self.stats['duplicates_skipped'] += 1
            return False
        test_case.id = ObjectId()
        test_case.created_by_id = self.user.id
        test_case.updated_by_id = self.user.id
        test_case.created_at = datetime.utcnow()
        test_case.updated_at = datetime.utcnow()
        self.existing_case_keys.add(case_key)
        self.pending_cases.append(test_case)
        self.pending_history.append(ChangeHistory(
            entity_type='test_case',
            entity_id=str(test_case.id),
            action='created',
            user_id=self.user.id,
            description=f'Imported test case: {test_case.title}'
        ))
        if test_case.steps:
            self.stats['steps_created'] += len(test_case.steps)
        if len(self.pending_cases) >= self.batch_size:
            self.flush()
        return True

    def flush(self):
        if not self.pending_cases:
            return
        try:
            TestCase.objects.insert(self.pending_cases, load_bulk=False)
            if self.pending_history:
                ChangeHistory.objects.insert(self.pending_history, load_bulk=False)
            self.stats['cases_created'] += len(self.pending_cases)
        except Exception as exc:
            self.stats['errors'].append(str(exc))
            # Fall back to per-document save to salvage work
            for case in self.pending_cases:
                try:
                    case.save()
                    ChangeHistory(
                        entity_type='test_case',
                        entity_id=str(case.id),
                        action='created',
                        user_id=self.user.id,
                        description=f'Imported test case: {case.title}'
                    ).save()
                    self.stats['cases_created'] += 1
                except Exception as inner_exc:
                    self.stats['errors'].append(str(inner_exc))
        finally:
            self.pending_cases = []
            self.pending_history = []

    def finalize(self):
        self.flush()
        return self.stats


def resolve_test_case(project, identifier, include_deleted=False):
    """Resolve either a full ObjectId or a friendly ID (e.g. C6922e056) to a real TestCase."""
    identifier = (identifier or '').strip()
    if not identifier or not project:
        return None
    
    # Normalized filters that apply to every lookup
    base_filters = {'project': project}
    if not include_deleted:
        base_filters['is_deleted'] = False
    
    # Try to extract a full 24-char ObjectId (ignore any surrounding text like "ID: <oid>")
    full_oid_match = re.search(r'([0-9a-fA-F]{24})', identifier)
    if full_oid_match:
        try:
            return TestCase.objects.get(
                id=ObjectId(full_oid_match.group(1)),
                **base_filters,
            )
        except (InvalidId, TestCase.DoesNotExist):
            # Fall through to friendly lookup
            pass
    
    # Support friendly IDs shown in UI (prefix C + first 8 hex chars, possibly followed by extra text)
    friendly_source = identifier.lower()
    if friendly_source.startswith('c'):
        friendly_source = friendly_source[1:]
    
    friendly_match = re.search(r'([0-9a-f]{8,24})', friendly_source)
    if not friendly_match:
        return None
    
    friendly = friendly_match.group(1)
    low_hex = friendly.ljust(24, '0')
    high_hex = friendly.ljust(24, 'f')
    
    try:
        low_id = ObjectId(low_hex)
        high_id = ObjectId(high_hex)
    except InvalidId:
        return None
    
    raw_filters = {
        '_id': {'$gte': low_id, '$lte': high_id},
        'project': project.id,
    }
    if not include_deleted:
        raw_filters['is_deleted'] = False
    
    return TestCase.objects(__raw__=raw_filters).first()


@login_required
def test_case_list(request, project_id=None):
    """List all test cases for a specific project, optionally filtered by section or test case."""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project - either from URL or query parameter
    project = None
    if project_id:
        try:
            project = Project.objects.get(id=ObjectId(project_id), is_active=True)
        except (Project.DoesNotExist, Exception):
            messages.error(request, 'Project not found.')
            return redirect('testcases:project_dashboard')
    else:
        project_id_param = request.GET.get('project')
        if project_id_param:
            try:
                project = Project.objects.get(id=ObjectId(project_id_param), is_active=True)
            except (Project.DoesNotExist, Exception):
                pass
    
    # If no project specified, redirect to project dashboard
    if not project:
        # Try to get the first active project as default
        default_project = Project.objects(is_active=True).first()
        if default_project:
            return redirect('testcases:test_case_list', project_id=str(default_project.id))
        else:
            messages.info(request, 'No projects available. Please create a project first.')
            return redirect('testcases:project_dashboard')
    
    # Get filter parameters
    section_filter = request.GET.get('section')
    test_case_filter = request.GET.get('test_case')
    sort_param = request.GET.get('sort', 'section')  # Default sort by section
    filter_param = request.GET.get('filter', 'all')  # Default filter: all
    search_query = request.GET.get('search', '').strip()  # Search query
    
    # If a specific test case is selected, get it for detail view
    selected_test_case = None
    if test_case_filter:
        selected_test_case = resolve_test_case(project, test_case_filter)
        if not selected_test_case:
            messages.error(request, f'Test case "{test_case_filter}" not found in this project.')
    
    # Query test cases - filter by project
    test_cases_query = TestCase.objects(project=project, is_deleted=False)
    
    # Apply search filter if provided
    matching_section_ids_for_cases = set()
    if search_query:
        # First, find all sections that match the search (within this project)
        all_sections_for_search = list(Section.objects(project=project).order_by('name'))
        for section in all_sections_for_search:
            if (search_query.lower() in section.name.lower() or 
                (section.description and search_query.lower() in section.description.lower())):
                matching_section_ids_for_cases.add(str(section.id))
        
        # Search in title, description, preconditions, AND include test cases from matching sections
        # Filter by title, description, preconditions containing search query OR section matches
        if matching_section_ids_for_cases:
            test_cases_query = test_cases_query.filter(
                __raw__={
                    '$or': [
                        {'title': {'$regex': search_query, '$options': 'i'}},
                        {'description': {'$regex': search_query, '$options': 'i'}},
                        {'preconditions': {'$regex': search_query, '$options': 'i'}},
                        {'section': {'$in': [ObjectId(sid) for sid in matching_section_ids_for_cases]}},
                    ]
                }
            )
        else:
            # No matching sections, only search in test case fields
            test_cases_query = test_cases_query.filter(
                __raw__={
                    '$or': [
                        {'title': {'$regex': search_query, '$options': 'i'}},
                        {'description': {'$regex': search_query, '$options': 'i'}},
                        {'preconditions': {'$regex': search_query, '$options': 'i'}},
                    ]
                }
            )
    
    # If a test case is selected, don't show the list - only show detail
    if selected_test_case:
        # When showing detail, we still need test_cases for context, but filter by section
        if selected_test_case.section:
            test_cases_query = test_cases_query.filter(section=selected_test_case.section)
    elif section_filter:
        # If section is selected, show that section's test cases AND test cases from its subsections
        try:
            section = Section.objects.get(project=project, id=ObjectId(section_filter))
            # Get all subsections recursively
            def get_all_subsection_ids(parent_section):
                """Recursively get all subsection IDs"""
                subsection_ids = []
                subsections = Section.objects(project=project, parent=parent_section)
                for subsection in subsections:
                    subsection_ids.append(subsection.id)
                    # Recursively get subsections of subsections
                    subsection_ids.extend(get_all_subsection_ids(subsection))
                return subsection_ids
            
            # Get all subsection IDs
            all_subsection_ids = get_all_subsection_ids(section)
            # Filter to show test cases in this section OR any of its subsections
            if all_subsection_ids:
                test_cases_query = test_cases_query.filter(
                    __raw__={
                        '$or': [
                            {'section': section.id},
                            {'section': {'$in': all_subsection_ids}}
                        ]
                    }
                )
            else:
                # No subsections, just show this section's test cases
                test_cases_query = test_cases_query.filter(section=section)
        except (Section.DoesNotExist, Exception):
            pass
    
    # Apply filter (only if not already filtered by section from tree selection)
    # Note: If section_filter is set, we already filtered by section above
    # So we only apply additional filters (type, priority) or override section filter if explicitly requested
    if filter_param == 'all':
        pass  # Show all (or what's already filtered by section)
    elif filter_param.startswith('section_'):
        # Filter by specific section (this overrides the tree section filter)
        try:
            filter_section_id = filter_param.replace('section_', '')
            filter_section = Section.objects.get(id=ObjectId(filter_section_id))
            # Override the section filter from tree
            test_cases_query = TestCase.objects(is_deleted=False).filter(section=filter_section)
        except (Section.DoesNotExist, Exception):
            pass
    elif filter_param.startswith('type_'):
        # Filter by type (applies on top of section filter if any)
        filter_type = filter_param.replace('type_', '')
        test_cases_query = test_cases_query.filter(type=filter_type)
    elif filter_param.startswith('priority_'):
        # Filter by priority (applies on top of section filter if any)
        filter_priority = filter_param.replace('priority_', '')
        test_cases_query = test_cases_query.filter(priority=filter_priority)
    
    # Apply sorting
    if sort_param == 'section':
        # Sort by section name, then by created date
        test_cases_list = list(test_cases_query)
        test_cases_list.sort(key=lambda tc: (
            tc.section.name if tc.section else '',
            tc.created_at
        ), reverse=True)
        test_cases = test_cases_list
    elif sort_param == 'title':
        test_cases = list(test_cases_query.order_by('title'))
    elif sort_param == 'date_newest':
        test_cases = list(test_cases_query.order_by('-created_at'))
    elif sort_param == 'date_oldest':
        test_cases = list(test_cases_query.order_by('created_at'))
    elif sort_param == 'type':
        test_cases = list(test_cases_query.order_by('type', '-created_at'))
    elif sort_param == 'priority':
        # Priority order: Critical, High, Medium, Low
        priority_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}
        test_cases_list = list(test_cases_query)
        test_cases_list.sort(key=lambda tc: (
            priority_order.get(tc.priority, 99),
            tc.created_at
        ))
        test_cases = test_cases_list
    else:
        test_cases = list(test_cases_query.order_by('-created_at'))
    
    # Build hierarchical tree structure from sections that have test cases (within this project)
    all_sections = list(Section.objects(project=project).order_by('name'))
    
    # Filter sections by search query if provided
    if search_query:
        # Find sections that match search OR have test cases that match
        matching_section_ids = set()
        # First, find sections that match by name/description
        for section in all_sections:
            if (search_query.lower() in section.name.lower() or 
                (section.description and search_query.lower() in section.description.lower())):
                matching_section_ids.add(str(section.id))
                # Also include all parent sections up to root
                current = section.parent
                while current:
                    matching_section_ids.add(str(current.id))
                    current = current.parent
                # Also include all child sections
                def add_children(section_obj):
                    children = Section.objects(parent=section_obj)
                    for child in children:
                        matching_section_ids.add(str(child.id))
                        add_children(child)  # Recursively add grandchildren
                add_children(section)
        
        # Also find sections that have matching test cases
        matching_test_cases = TestCase.objects(
            is_deleted=False,
            __raw__={
                '$or': [
                    {'title': {'$regex': search_query, '$options': 'i'}},
                    {'description': {'$regex': search_query, '$options': 'i'}},
                    {'preconditions': {'$regex': search_query, '$options': 'i'}},
                ]
            }
        )
        for test_case in matching_test_cases:
            if test_case.section:
                matching_section_ids.add(str(test_case.section.id))
                # Include parent sections
                current = test_case.section.parent
                while current:
                    matching_section_ids.add(str(current.id))
                    current = current.parent
        
        # Filter sections to only include matching ones or their parents
        if matching_section_ids:
            all_sections = [s for s in all_sections if str(s.id) in matching_section_ids]
        else:
            # If no sections match, show empty list
            all_sections = []
    
    # Build tree structure
    def build_section_tree(sections, parent=None):
        """Build hierarchical tree of sections"""
        tree = []
        for section in sections:
            # Check if this section belongs to the current parent level
            section_has_parent = section.parent is not None
            parent_is_none = parent is None
            
            # Match logic:
            # - If parent is None, include sections with no parent
            # - If parent is set, include sections where section.parent == parent
            if parent_is_none and not section_has_parent:
                # Root level section (no parent)
                # Get test cases for this section
                if search_query:
                    # Filter test cases by search query
                    section_test_cases_query = TestCase.objects(section=section, is_deleted=False)
                    # Check if section matches search
                    section_matches = (search_query.lower() in section.name.lower() or 
                                      (section.description and search_query.lower() in section.description.lower()))
                    if section_matches:
                        # If section matches, show all test cases in it
                        section_test_cases = list(section_test_cases_query.order_by('id'))
                    else:
                        # If section doesn't match, only show matching test cases
                        section_test_cases = list(section_test_cases_query.filter(
                            __raw__={
                                '$or': [
                                    {'title': {'$regex': search_query, '$options': 'i'}},
                                    {'description': {'$regex': search_query, '$options': 'i'}},
                                    {'preconditions': {'$regex': search_query, '$options': 'i'}},
                                ]
                            }
                        ).order_by('id'))
                else:
                    section_test_cases = list(TestCase.objects(section=section, is_deleted=False).order_by('id'))
                
                # Check if this section has any subsections in the database
                direct_subsections_count = Section.objects(parent=section).count()
                
                # Recursively build subsections
                subsections = build_section_tree(sections, parent=section)
                has_test_cases = len(section_test_cases) > 0
                has_subsections = len(subsections) > 0
                has_direct_subsections = direct_subsections_count > 0
                
                # ALWAYS include root sections - they should appear in the tree even if empty
                # This ensures newly created sections are visible
                section_matches_search = search_query and (search_query.lower() in section.name.lower() or 
                                                          (section.description and search_query.lower() in section.description.lower()))
                tree.append({
                    'section': section,
                    'test_cases': section_test_cases,
                    'subsections': subsections,
                    'has_cases': has_test_cases or has_subsections or has_direct_subsections,
                    'matches_search': section_matches_search if search_query else False,
                })
            elif not parent_is_none and section_has_parent and str(section.parent.id) == str(parent.id):
                # Subsection (has parent matching current parent)
                # Get test cases for this section
                if search_query:
                    # Filter test cases by search query
                    section_test_cases_query = TestCase.objects(section=section, is_deleted=False)
                    # Check if section matches search
                    section_matches = (search_query.lower() in section.name.lower() or 
                                      (section.description and search_query.lower() in section.description.lower()))
                    if section_matches:
                        # If section matches, show all test cases in it
                        section_test_cases = list(section_test_cases_query.order_by('id'))
                    else:
                        # If section doesn't match, only show matching test cases
                        section_test_cases = list(section_test_cases_query.filter(
                            __raw__={
                                '$or': [
                                    {'title': {'$regex': search_query, '$options': 'i'}},
                                    {'description': {'$regex': search_query, '$options': 'i'}},
                                    {'preconditions': {'$regex': search_query, '$options': 'i'}},
                                ]
                            }
                        ).order_by('id'))
                else:
                    section_test_cases = list(TestCase.objects(section=section, is_deleted=False).order_by('id'))
                
                # Check if this section has any subsections in the database
                direct_subsections_count = Section.objects(parent=section).count()
                
                # Recursively build subsections
                subsections = build_section_tree(sections, parent=section)
                has_test_cases = len(section_test_cases) > 0
                has_subsections = len(subsections) > 0
                has_direct_subsections = direct_subsections_count > 0
                
                # ALWAYS include subsections - they should appear in the tree even if empty
                # This ensures newly created subsections are visible
                section_matches_search = search_query and (search_query.lower() in section.name.lower() or 
                                                          (section.description and search_query.lower() in section.description.lower()))
                tree.append({
                    'section': section,
                    'test_cases': section_test_cases,
                    'subsections': subsections,
                    'has_cases': has_test_cases or has_subsections or has_direct_subsections,
                    'matches_search': section_matches_search if search_query else False,
                })
        return tree
    
    # Build tree starting from root sections (no parent)
    section_tree = build_section_tree(all_sections, parent=None)
    
    # Get sections that have test cases for filter dropdown
    sections_with_cases = []
    for section in all_sections:
        if TestCase.objects(section=section, is_deleted=False).count() > 0:
            sections_with_cases.append(section)
    
    # Count total sections and cases (only those that appear in tree)
    def count_sections_in_tree(tree):
        """Recursively count all sections in the tree"""
        count = 0
        for item in tree:
            count += 1  # Count this section
            count += count_sections_in_tree(item['subsections'])  # Count subsections
        return count
    
    total_sections = count_sections_in_tree(section_tree)
    total_cases = TestCase.objects(is_deleted=False).count()
    
    # Get steps for selected test case if one is selected
    selected_test_case_steps = None
    selected_test_case_history = None
    if selected_test_case:
        selected_test_case_steps = sorted(selected_test_case.steps, key=lambda x: x.step_number) if selected_test_case.steps else []
        # Get change history for selected test case
        selected_test_case_history = ChangeHistory.objects(
            entity_type='test_case',
            entity_id=str(selected_test_case.id)
        ).order_by('-timestamp')[:50]  # Limit to last 50 changes
    
    # Get selected section object if section is selected
    selected_section_obj = None
    selected_section_subsections = []
    if section_filter:
        try:
            selected_section_obj = Section.objects.get(project=project, id=ObjectId(section_filter))
            # Get all direct subsections of the selected section with their test case counts
            subsections = Section.objects(project=project, parent=selected_section_obj).order_by('name')
            selected_section_subsections = []
            for subsection in subsections:
                # Count test cases in this subsection
                tc_count = TestCase.objects(project=project, section=subsection, is_deleted=False).count()
                selected_section_subsections.append({
                    'subsection': subsection,
                    'test_case_count': tc_count
                })
        except (Section.DoesNotExist, Exception):
            pass
    
    # Get unique types and priorities for filter dropdown
    all_test_cases = TestCase.objects(is_deleted=False)
    unique_types = sorted(set(tc.type for tc in all_test_cases if tc.type))
    unique_priorities = sorted(set(tc.priority for tc in all_test_cases if tc.priority))
    
    context = {
        'project': project,
        'test_cases': test_cases,
        'sections': sections_with_cases,
        'all_sections': all_sections,
        'unique_types': unique_types,
        'unique_priorities': unique_priorities,
        'selected_section': section_filter,
        'selected_section_obj': selected_section_obj,
        'selected_section_subsections': selected_section_subsections,
        'selected_test_case': test_case_filter,
        'selected_test_case_obj': selected_test_case,
        'selected_test_case_steps': selected_test_case_steps,
        'selected_test_case_history': selected_test_case_history,
        'section_tree': section_tree,
        'total_sections': total_sections,
        'total_cases': total_cases,
        'current_sort': sort_param,
        'current_filter': filter_param,
        'search_query': search_query,
        'user_can_edit': user_can_edit_sections(request.user),
        'user_can_delete': user_can_delete_sections(request.user),
    }
    return render(request, 'testcases/test_case_list.html', context)


@login_required
def test_case_add(request, project_id):
    """Add a new test case"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    try:
        project = Project.objects.get(id=ObjectId(project_id), is_active=True)
    except (Project.DoesNotExist, Exception):
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    if request.method == 'POST':
        form = TestCaseForm(request.POST)
        if form.is_valid():
            # Get or create section
            section_id = form.cleaned_data.get('section')
            new_section_name = form.cleaned_data.get('new_section')
            
            if new_section_name:
                # Create new section (mongoengine doesn't have get_or_create)
                section_name = new_section_name.strip()
                try:
                    section = Section.objects.get(project=project, name=section_name)
                except Section.DoesNotExist:
                    section = Section(project=project, name=section_name, created_by_id=request.user.id)
                    section.save()
            elif section_id:
                try:
                    section = Section.objects.get(project=project, id=ObjectId(section_id))
                except (Section.DoesNotExist, Exception):
                    section = Section(project=project, name='Default', created_by_id=request.user.id)
                    section.save()
            else:
                section = Section(project=project, name='Default', created_by_id=request.user.id)
                section.save()
            
            # Create test case
            test_case = TestCase(
                project=project,
                title=form.cleaned_data['title'],
                section=section,
                type=form.cleaned_data['type'],
                priority=form.cleaned_data['priority'],
                estimate=form.cleaned_data.get('estimate', ''),
                automation_type=form.cleaned_data.get('automation_type', 'None'),
                labels=form.cleaned_data.get('labels', ''),
                description=form.cleaned_data.get('description', ''),
                preconditions=form.cleaned_data.get('preconditions', ''),
                template='Test Case (Steps)',  # Always set to this
                created_by_id=request.user.id,
                is_deleted=False
            )
            
            # Handle test steps - steps are required
            steps_data = request.POST.getlist('steps[]')
            expected_results = request.POST.getlist('expected_results[]')
            
            if not steps_data or not any(step.strip() for step in steps_data):
                form.add_error(None, "At least one test step with description is required.")
                context = {'form': form}
                return render(request, 'testcases/test_case_add.html', context)
            
            test_steps = []
            for idx, (step_desc, expected) in enumerate(zip(steps_data, expected_results), start=1):
                if step_desc.strip():  # Only create step if description is not empty
                    if not expected.strip():
                        form.add_error(None, f"Step {idx} requires an expected result.")
                        context = {'form': form}
                        return render(request, 'testcases/test_case_add.html', context)
                    test_steps.append(TestStep(
                        step_number=idx,
                        description=step_desc.strip(),
                        expected_result=expected.strip()
                    ))
            
            if not test_steps:
                form.add_error(None, "At least one test step with description and expected result is required.")
                context = {'form': form}
                return render(request, 'testcases/test_case_add.html', context)
            
            test_case.steps = test_steps
            test_case.save()
            
            # Record change history
            record_change('test_case', str(test_case.id), 'created', request.user.id, f'Created test case: {test_case.title}')
            
            messages.success(request, f'Test case "{test_case.title}" created successfully!')
            
            # Check if "Add & Next" was clicked
            if request.POST.get('add_and_next'):
                return redirect('testcases:test_case_add', project_id=str(project.id))
            return redirect('testcases:test_case_list', project_id=str(project.id))
    else:
        form = TestCaseForm()
        # Filter sections by project
        form.fields['section'].queryset = Section.objects(project=project).order_by('name')
    
    # Pre-select section if provided in query params
    initial_section = request.GET.get('section')
    if initial_section:
        form.fields['section'].initial = initial_section
    
    context = {
        'project': project,
        'form': form,
    }
    return render(request, 'testcases/test_case_add.html', context)


@login_required
def test_case_detail(request, project_id, pk):
    """View test case details"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    test_case = resolve_test_case(project, pk)
    if not test_case:
        raise Http404("Test case not found")
    
    # Steps are embedded, so we can access them directly
    steps = sorted(test_case.steps, key=lambda x: x.step_number) if test_case.steps else []
    
    # Get change history for this test case
    change_history = ChangeHistory.objects(
        entity_type='test_case',
        entity_id=str(test_case.id)
    ).order_by('-timestamp')[:50]  # Limit to last 50 changes
    
    context = {
        'project': project,
        'test_case': test_case,
        'steps': steps,
        'user_can_edit': user_can_edit_sections(request.user),
        'user_can_delete': user_can_delete_sections(request.user),
        'change_history': change_history,
    }
    return render(request, 'testcases/test_case_detail.html', context)


@login_required
def test_case_edit(request, project_id, pk):
    """Edit an existing test case"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Check permissions
    if not user_can_edit_sections(request.user):
        messages.error(request, 'You do not have permission to edit test cases.')
        try:
            test_case = resolve_test_case(project, pk)
            return redirect('testcases:test_case_detail', project_id=str(project.id), pk=str(test_case.id))
        except Exception:
            return redirect('testcases:test_case_list', project_id=str(project.id))
    
    test_case = resolve_test_case(project, pk)
    if not test_case:
        raise Http404("Test case not found")
    
    if request.method == 'POST':
        form = TestCaseForm(request.POST)
        if form.is_valid():
            # Get or create section
            section_id = form.cleaned_data.get('section')
            new_section_name = form.cleaned_data.get('new_section')
            
            if new_section_name:
                # Create new section (mongoengine doesn't have get_or_create)
                section_name = new_section_name.strip()
                try:
                    section = Section.objects.get(project=project, name=section_name)
                except Section.DoesNotExist:
                    section = Section(project=project, name=section_name, created_by_id=request.user.id)
                    section.save()
            elif section_id:
                try:
                    section = Section.objects.get(project=project, id=ObjectId(section_id))
                except (Section.DoesNotExist, Exception):
                    section = test_case.section
            else:
                section = test_case.section
            
            # Update test case
            test_case.title = form.cleaned_data['title']
            test_case.section = section
            test_case.type = form.cleaned_data['type']
            test_case.priority = form.cleaned_data['priority']
            test_case.estimate = form.cleaned_data.get('estimate', '')
            test_case.automation_type = form.cleaned_data.get('automation_type', 'None')
            test_case.labels = form.cleaned_data.get('labels', '')
            test_case.description = form.cleaned_data.get('description', '')
            test_case.preconditions = form.cleaned_data.get('preconditions', '')
            
            # Handle test steps - steps are required
            steps_data = request.POST.getlist('steps[]')
            expected_results = request.POST.getlist('expected_results[]')
            
            if not steps_data or not any(step.strip() for step in steps_data):
                form.add_error(None, "At least one test step with description is required.")
                existing_steps = sorted(test_case.steps, key=lambda x: x.step_number) if test_case.steps else []
                context = {
                    'form': form,
                    'test_case': test_case,
                    'existing_steps': existing_steps,
                }
                return render(request, 'testcases/test_case_edit.html', context)
            
            test_steps = []
            for idx, (step_desc, expected) in enumerate(zip(steps_data, expected_results), start=1):
                if step_desc.strip():
                    if not expected.strip():
                        form.add_error(None, f"Step {idx} requires an expected result.")
                        existing_steps = sorted(test_case.steps, key=lambda x: x.step_number) if test_case.steps else []
                        context = {
                            'form': form,
                            'test_case': test_case,
                            'existing_steps': existing_steps,
                        }
                        return render(request, 'testcases/test_case_edit.html', context)
                    test_steps.append(TestStep(
                        step_number=idx,
                        description=step_desc.strip(),
                        expected_result=expected.strip()
                    ))
            
            if not test_steps:
                form.add_error(None, "At least one test step with description and expected result is required.")
                existing_steps = sorted(test_case.steps, key=lambda x: x.step_number) if test_case.steps else []
                context = {
                    'form': form,
                    'test_case': test_case,
                    'existing_steps': existing_steps,
                }
                return render(request, 'testcases/test_case_edit.html', context)
            
            test_case.steps = test_steps
            test_case.updated_by_id = request.user.id
            test_case.save()
            
            # Record change history
            record_change('test_case', str(test_case.id), 'updated', request.user.id, f'Updated test case: {test_case.title}')
            
            messages.success(request, f'Test case "{test_case.title}" updated successfully!')
            return redirect('testcases:test_case_detail', project_id=str(project.id), pk=str(test_case.id))
    else:
        # Pre-populate form with existing data
        form = TestCaseForm(initial={
            'title': test_case.title,
            'section': str(test_case.section.id) if test_case.section else '',
            'type': test_case.type,
            'priority': test_case.priority,
            'estimate': test_case.estimate,
            'automation_type': test_case.automation_type,
            'labels': test_case.labels,
            'description': getattr(test_case, 'description', ''),
            'preconditions': test_case.preconditions,
        })
        # Filter sections by project
        form.fields['section'].queryset = Section.objects(project=project).order_by('name')
    
    existing_steps = sorted(test_case.steps, key=lambda x: x.step_number) if test_case.steps else []
    
    context = {
        'project': project,
        'form': form,
        'test_case': test_case,
        'existing_steps': existing_steps,
    }
    return render(request, 'testcases/test_case_edit.html', context)


@login_required
@require_http_methods(["POST"])
def test_case_delete(request, project_id, pk):
    """Delete a test case (soft delete)"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Check permissions
    if not user_can_delete_sections(request.user):
        messages.error(request, 'You do not have permission to delete test cases.')
        return redirect('testcases:test_case_list', project_id=str(project.id))
    
    test_case = resolve_test_case(project, pk)
    if test_case:
        test_case.is_deleted = True
        test_case.updated_by_id = request.user.id
        test_case.save()
        
        # Record change history
        record_change('test_case', str(test_case.id), 'deleted', request.user.id, f'Deleted test case: {test_case.title}')
        
        messages.success(request, f'Test case "{test_case.title}" deleted successfully!')
    else:
        messages.error(request, 'Test case not found.')
    
    return redirect('testcases:test_case_list', project_id=str(project.id))


@login_required
@require_http_methods(["POST"])
def section_add(request, project_id):
    """Add a new root section (no parent)"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    try:
        project = Project.objects.get(id=ObjectId(project_id), is_active=True)
    except (Project.DoesNotExist, Exception):
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    section_name = request.POST.get('section_name', '').strip()
    section_description = request.POST.get('section_description', '').strip()
    
    if section_name:
        try:
            # Check if section already exists at root level (no parent) in this project
            try:
                existing = Section.objects.get(project=project, name=section_name, parent__exists=False)
                messages.warning(request, f'Section "{section_name}" already exists.')
            except Section.DoesNotExist:
                try:
                    # Also try checking with parent=None
                    existing = Section.objects.get(project=project, name=section_name, parent=None)
                    messages.warning(request, f'Section "{section_name}" already exists.')
                except Section.DoesNotExist:
                    # Create new root section
                    try:
                        section = Section(
                            project=project,
                            name=section_name,
                            parent=None,
                            description=section_description or '',
                            created_by_id=request.user.id
                        )
                        # Validate before saving
                        section.clean()
                        section.save()
                        
                        # Record change history
                        record_change('section', str(section.id), 'created', request.user.id, f'Created section: {section_name}')
                        
                        # Verify it was saved
                        saved_section = Section.objects.get(id=section.id)
                        messages.success(request, f'Section "{section_name}" created successfully!')
                    except Exception as e:
                        import traceback
                        error_details = traceback.format_exc()
                        messages.error(request, f'Error creating section: {str(e)}')
                        # Log the full error for debugging
                        print(f"Error creating section: {error_details}")
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    else:
        messages.error(request, 'Section name is required.')
    
    # Redirect back to test case list
    return redirect('testcases:test_case_list', project_id=str(project.id))


@login_required
def section_add_subsection(request, project_id, section_id):
    """Add a subsection to an existing section"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    if request.method == 'POST':
        subsection_name = request.POST.get('subsection_name', '').strip()
        subsection_description = request.POST.get('subsection_description', '').strip()
        
        if subsection_name:
            try:
                parent_section = Section.objects.get(project=project, id=ObjectId(section_id))
                # Check if subsection already exists under this parent
                try:
                    existing = Section.objects.get(project=project, name=subsection_name, parent=parent_section)
                    messages.warning(request, f'Subsection "{subsection_name}" already exists under "{parent_section.name}".')
                except Section.DoesNotExist:
                    # Create new subsection
                    try:
                        subsection = Section(
                            project=project,
                            name=subsection_name,
                            parent=parent_section,
                            description=subsection_description if subsection_description else '',
                            created_by_id=request.user.id
                        )
                        # Validate before saving
                        subsection.clean()
                        subsection.save()
                        
                        # Record change history
                        record_change('section', str(subsection.id), 'created', request.user.id, f'Created subsection: {subsection_name} under {parent_section.name}')
                        
                        # Verify it was saved
                        saved_subsection = Section.objects.get(id=subsection.id)
                        messages.success(request, f'Subsection "{subsection_name}" created successfully under "{parent_section.name}"!')
                    except Exception as e:
                        import traceback
                        error_details = traceback.format_exc()
                        messages.error(request, f'Error creating subsection: {str(e)}')
                        # Log the full error for debugging
                        print(f"Error creating subsection: {error_details}")
            except (Section.DoesNotExist, Exception) as e:
                messages.error(request, f'Error: {str(e)}')
        else:
            messages.error(request, 'Subsection name is required.')
        
        # Redirect with query parameter
        return redirect(f"{reverse('testcases:test_case_list', args=[str(project.id)])}?section={section_id}")
    else:
        # GET request - redirect back
        return redirect(f"{reverse('testcases:test_case_list', args=[str(project.id)])}?section={section_id}")


@login_required
def section_manage(request, project_id):
    """Manage sections - list, edit, and delete sections"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Handle permission updates (admin only) - check before other POST handlers
    if request.method == 'POST' and 'update_permissions' in request.POST:
        if request.user.role != 3:  # Only admins
            messages.error(request, 'Only administrators can manage permissions.')
            return redirect('testcases:section_manage', project_id=str(project.id))
        
        user_id = request.POST.get('user_id')
        can_edit = request.POST.get('can_edit') == 'on'
        can_delete = request.POST.get('can_delete') == 'on'
        
        if user_id:
            try:
                # MongoEngine doesn't have get_or_create, so we need to do it manually
                try:
                    permission = SectionPermission.objects.get(user_id=int(user_id))
                except SectionPermission.DoesNotExist:
                    permission = SectionPermission(user_id=int(user_id))
                
                permission.can_edit = can_edit
                permission.can_delete = can_delete
                permission.save()
                messages.success(request, f'Permissions updated successfully.')
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                messages.error(request, f'Error updating permissions: {str(e)}')
                print(f"Error updating permissions: {error_details}")
        return redirect('testcases:section_manage', project_id=str(project.id))
    
    # Handle bulk delete of sections
    if request.method == 'POST' and 'delete_sections' in request.POST:
        # Check permissions
        if not user_can_delete_sections(request.user):
            messages.error(request, 'You do not have permission to delete sections.')
            return redirect('testcases:section_manage', project_id=str(project.id))
        section_ids = request.POST.getlist('section_ids')
        deleted_count = 0
        
        # Helper function to recursively delete subsections
        def delete_subsection_recursive(subsection):
            # Delete test cases in subsection (soft delete)
            subsection_test_cases = TestCase.objects(project=project, section=subsection, is_deleted=False)
            for test_case in subsection_test_cases:
                test_case.is_deleted = True
                test_case.updated_by_id = request.user.id
                test_case.save()
                # Record change history
                record_change('test_case', str(test_case.id), 'deleted', request.user.id, f'Deleted test case: {test_case.title} (via bulk section deletion)')
            
            # Recursively delete child subsections
            child_subsections = Section.objects(project=project, parent=subsection)
            for child in child_subsections:
                delete_subsection_recursive(child)
            
            # Record change history before deleting subsection
            record_change('section', str(subsection.id), 'deleted', request.user.id, f'Deleted subsection: {subsection.name} (via bulk deletion)')
            # Delete the subsection itself
            subsection.delete()
        
        for section_id in section_ids:
            try:
                section = Section.objects.get(project=project, id=ObjectId(section_id))
                section_name = section.name
                
                # Delete all test cases in this section (soft delete)
                test_cases = TestCase.objects(project=project, section=section, is_deleted=False)
                for test_case in test_cases:
                    test_case.is_deleted = True
                    test_case.updated_by_id = request.user.id
                    test_case.save()
                    # Record change history
                    record_change('test_case', str(test_case.id), 'deleted', request.user.id, f'Deleted test case: {test_case.title} (via bulk section deletion)')
                
                # Delete all subsections recursively
                subsections = Section.objects(project=project, parent=section)
                for subsection in subsections:
                    delete_subsection_recursive(subsection)
                
                # Record change history before deleting section
                record_change('section', str(section.id), 'deleted', request.user.id, f'Deleted section: {section_name} (via bulk deletion)')
                # Delete the section itself
                section.delete()
                deleted_count += 1
            except (Section.DoesNotExist, Exception) as e:
                pass
        
        if deleted_count > 0:
            messages.success(request, f'Successfully deleted {deleted_count} section(s) and all their test cases and subsections.')
        else:
            messages.warning(request, 'No sections were deleted.')
        return redirect('testcases:section_manage', project_id=str(project.id))
    
    # Handle bulk delete of test cases
    if request.method == 'POST' and 'delete_test_cases' in request.POST:
        test_case_ids = request.POST.getlist('test_case_ids')
        deleted_count = 0
        for test_case_id in test_case_ids:
            try:
                test_case = TestCase.objects.get(project=project, id=ObjectId(test_case_id), is_deleted=False)
                test_case.is_deleted = True
                test_case.updated_by_id = request.user.id
                test_case.save()
                # Record change history
                record_change('test_case', str(test_case.id), 'deleted', request.user.id, f'Deleted test case: {test_case.title} (via bulk deletion)')
                deleted_count += 1
            except (TestCase.DoesNotExist, Exception):
                pass
        
        if deleted_count > 0:
            messages.success(request, f'Successfully deleted {deleted_count} test case(s).')
        else:
            messages.warning(request, 'No test cases were deleted.')
        return redirect('testcases:section_manage', project_id=str(project.id))
    
    # Get all sections with their hierarchy info (filtered by project)
    all_sections = list(Section.objects(project=project).order_by('name'))
    
    # Build section data with counts and test cases
    sections_data = []
    all_test_cases = []
    
    for section in all_sections:
        # Get test cases in this section
        test_cases = list(TestCase.objects(section=section, is_deleted=False).order_by('-created_at'))
        
        # Count test cases in this section
        test_case_count = len(test_cases)
        
        # Count subsections
        subsection_count = Section.objects(project=project, parent=section).count()
        
        sections_data.append({
            'section': section,
            'test_case_count': test_case_count,
            'subsection_count': subsection_count,
            'parent_name': section.parent.name if section.parent else None,
            'test_cases': test_cases,  # Include test cases for display
        })
        
        # Add to all test cases list
        all_test_cases.extend(test_cases)
    
    # Get all users for permission management (admin only)
    # Only load this data if user is admin to prevent unnecessary queries
    all_users = []
    permissions_data = []
    is_admin = request.user.role == 3
    
    if is_admin:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        all_users = list(User.objects.filter(is_active=True).order_by('username'))
        
        # Get permissions for all users (only for admins)
        for user in all_users:
            try:
                permission = SectionPermission.objects.get(user_id=user.id)
                permissions_data.append({
                    'user': user,
                    'permission': permission,
                    'can_edit': permission.can_edit,
                    'can_delete': permission.can_delete,
                })
            except SectionPermission.DoesNotExist:
                permissions_data.append({
                    'user': user,
                    'permission': None,
                    'can_edit': False,
                    'can_delete': False,
                })
    
    context = {
        'project': project,
        'sections': sections_data,
        'all_test_cases': all_test_cases,
        'all_users': all_users,  # Empty list for non-admins
        'permissions_data': permissions_data,  # Empty list for non-admins
        'user_can_edit': user_can_edit_sections(request.user),
        'user_can_delete': user_can_delete_sections(request.user),
        'is_admin': is_admin,  # Explicitly set for clarity
    }
    return render(request, 'testcases/section_manage.html', context)


@login_required
def section_edit(request, project_id, section_id):
    """Edit a section"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Check permissions
    if not user_can_edit_sections(request.user):
        messages.error(request, 'You do not have permission to edit sections.')
        return redirect('testcases:section_manage', project_id=str(project.id))
    
    try:
        section = Section.objects.get(project=project, id=ObjectId(section_id))
    except (Section.DoesNotExist, Exception):
        messages.error(request, 'Section not found.')
        return redirect('testcases:section_manage', project_id=str(project.id))
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        
        if name:
            # Check if name already exists (excluding current section) - within same project
            try:
                existing = Section.objects.get(project=project, name=name, parent=section.parent)
                if str(existing.id) != str(section.id):
                    messages.error(request, f'A section with name "{name}" already exists in this parent.')
                    return redirect('testcases:section_edit', project_id=str(project.id), section_id=section_id)
            except Section.DoesNotExist:
                pass
            
            section.name = name
            section.description = description if description else ''
            section.updated_by_id = request.user.id
            section.save()
            
            # Record change history
            record_change('section', str(section.id), 'updated', request.user.id, f'Updated section: {section.name}')
            
            messages.success(request, f'Section "{section.name}" updated successfully!')
            return redirect('testcases:section_manage', project_id=str(project.id))
        else:
            messages.error(request, 'Section name is required.')
    
    # Get all sections for parent selection (excluding current section and its descendants) - within same project
    available_parents = []
    for s in Section.objects(project=project).order_by('name'):
        if str(s.id) != str(section.id) and (not section.parent or str(s.id) != str(section.parent.id)):
            # Check if section is a descendant of current section (to prevent circular references)
            is_descendant = False
            current = section.parent
            while current:
                if str(current.id) == str(s.id):
                    is_descendant = True
                    break
                current = current.parent
            
            if not is_descendant:
                available_parents.append(s)
    
    context = {
        'project': project,
        'section': section,
        'available_parents': available_parents,
    }
    return render(request, 'testcases/section_edit.html', context)


@login_required
@require_http_methods(["POST"])
def section_delete(request, project_id, section_id):
    """Delete a section and all its test cases and subsections"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Check permissions
    if not user_can_delete_sections(request.user):
        messages.error(request, 'You do not have permission to delete sections.')
        return redirect('testcases:section_manage', project_id=str(project.id))
    
    try:
        section = Section.objects.get(project=project, id=ObjectId(section_id))
    except (Section.DoesNotExist, Exception):
        messages.error(request, 'Section not found.')
        return redirect('testcases:section_manage', project_id=str(project.id))
    
    section_name = section.name
    
    # Delete all test cases in this section (soft delete)
    test_cases = TestCase.objects(project=project, section=section, is_deleted=False)
    test_case_count = test_cases.count()
    for test_case in test_cases:
        test_case.is_deleted = True
        test_case.updated_by_id = request.user.id
        test_case.save()
        # Record change history
        record_change('test_case', str(test_case.id), 'deleted', request.user.id, f'Deleted test case: {test_case.title} (via section deletion)')
    
    # Delete all subsections recursively
    def delete_subsection_recursive(subsection):
        # Delete test cases in subsection
        subsection_test_cases = TestCase.objects(project=project, section=subsection, is_deleted=False)
        for test_case in subsection_test_cases:
            test_case.is_deleted = True
            test_case.updated_by_id = request.user.id
            test_case.save()
            # Record change history
            record_change('test_case', str(test_case.id), 'deleted', request.user.id, f'Deleted test case: {test_case.title} (via section deletion)')
        
        # Recursively delete child subsections
        child_subsections = Section.objects(project=project, parent=subsection)
        for child in child_subsections:
            delete_subsection_recursive(child)
        
        # Record change history before deleting subsection
        record_change('section', str(subsection.id), 'deleted', request.user.id, f'Deleted subsection: {subsection.name}')
        # Delete the subsection itself
        subsection.delete()
    
    subsections = Section.objects(project=project, parent=section)
    subsection_count = subsections.count()
    for subsection in subsections:
        delete_subsection_recursive(subsection)
    
    # Record change history before deleting section
    record_change('section', str(section.id), 'deleted', request.user.id, f'Deleted section: {section_name}')
    # Delete the section itself
    section.delete()
    
    messages.success(request, f'Section "{section_name}" and all its {test_case_count} test case(s) and {subsection_count} subsection(s) deleted successfully!')
    return redirect('testcases:section_manage', project_id=str(project.id))


@login_required
def test_case_export(request, project_id):
    """Export test cases to Excel or CSV format"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    format_type = request.GET.get('format', 'excel')  # 'excel' or 'csv'
    
    # Get all non-deleted test cases for this project
    test_cases = TestCase.objects(project=project, is_deleted=False).order_by('section__name', 'title')
    
    if format_type == 'csv':
        # Create CSV response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="test_cases_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'ID', 'Title', 'Section', 'Type', 'Priority', 'Estimate', 
            'Automation Type', 'Labels', 'Description', 'Preconditions',
            'Step 1', 'Expected Result 1', 'Step 2', 'Expected Result 2',
            'Step 3', 'Expected Result 3', 'Step 4', 'Expected Result 4',
            'Step 5', 'Expected Result 5', 'Created At', 'Updated At'
        ])
        
        # Write test cases
        for test_case in test_cases:
            # Get section path (parent sections)
            section_path = []
            current_section = test_case.section
            while current_section:
                section_path.insert(0, current_section.name)
                current_section = current_section.parent
            section_full_path = ' > '.join(section_path)
            
            # Prepare steps data (up to 5 steps)
            steps_data = []
            for i, step in enumerate(sorted(test_case.steps, key=lambda x: x.step_number)[:5], 1):
                steps_data.extend([step.description, step.expected_result])
            
            # Pad steps if less than 5
            while len(steps_data) < 10:
                steps_data.extend(['', ''])
            
            writer.writerow([
                str(test_case.id)[:8],  # ID (first 8 chars)
                test_case.title,
                section_full_path,
                test_case.type,
                test_case.priority,
                test_case.estimate or '',
                test_case.automation_type,
                test_case.labels or '',
                test_case.description or '',
                test_case.preconditions or '',
                *steps_data,
                test_case.created_at.strftime('%Y-%m-%d %H:%M:%S') if test_case.created_at else '',
                test_case.updated_at.strftime('%Y-%m-%d %H:%M:%S') if test_case.updated_at else '',
            ])
        
        return response
    
    else:  # Excel format
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = [
            'ID', 'Title', 'Section', 'Type', 'Priority', 'Estimate', 
            'Automation Type', 'Labels', 'Description', 'Preconditions',
            'Step 1', 'Expected Result 1', 'Step 2', 'Expected Result 2',
            'Step 3', 'Expected Result 3', 'Step 4', 'Expected Result 4',
            'Step 5', 'Expected Result 5', 'Created At', 'Updated At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write test cases
        for row_num, test_case in enumerate(test_cases, 2):
            # Get section path (parent sections)
            section_path = []
            current_section = test_case.section
            while current_section:
                section_path.insert(0, current_section.name)
                current_section = current_section.parent
            section_full_path = ' > '.join(section_path)
            
            # Prepare steps data (up to 5 steps)
            steps_data = []
            for i, step in enumerate(sorted(test_case.steps, key=lambda x: x.step_number)[:5], 1):
                steps_data.extend([step.description, step.expected_result])
            
            # Pad steps if less than 5
            while len(steps_data) < 10:
                steps_data.extend(['', ''])
            
            row_data = [
                str(test_case.id)[:8],  # ID (first 8 chars)
                test_case.title,
                section_full_path,
                test_case.type,
                test_case.priority,
                test_case.estimate or '',
                test_case.automation_type,
                test_case.labels or '',
                test_case.description or '',
                test_case.preconditions or '',
                *steps_data,
                test_case.created_at.strftime('%Y-%m-%d %H:%M:%S') if test_case.created_at else '',
                test_case.updated_at.strftime('%Y-%m-%d %H:%M:%S') if test_case.updated_at else '',
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="test_cases_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response


@login_required
def test_case_import(request, project_id):
    """Import test cases from Excel or CSV file"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get project
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    if not user_can_edit_sections(request.user):
        messages.error(request, 'You do not have permission to import test cases.')
        return redirect('testcases:test_case_list', project_id=str(project.id))
    
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'Please select a file to import.')
            return redirect('testcases:test_case_list', project_id=str(project.id))
        
        # Ensure indexes are correct before importing sections
        try:
            ensure_section_indexes()
        except Exception as index_error:
            messages.error(request, f'Failed to prepare section indexes: {index_error}')
            return redirect('testcases:test_case_list', project_id=str(project.id))
        
        importer = TestCaseImportService(project, request.user)
        
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        
        imported_count = 0
        error_count = 0
        errors = []
        
        try:
            if file_name.endswith('.csv'):
                # Process CSV file
                decoded_file = uploaded_file.read().decode('utf-8-sig')
                csv_reader = csv.DictReader(io.StringIO(decoded_file))
                
                for row_num, row in enumerate(csv_reader, start=2):
                    importer.stats['rows_processed'] += 1
                    try:
                        # Get or create section - try multiple column name variations
                        # Priority: Section Hie (Section Hierarchy) > Section > Section Des > Suite
                        section_path = (
                            row.get('Section Hie', '').strip() or
                            row.get('Section Hier', '').strip() or
                            row.get('Section Hierarchical', '').strip() or
                            row.get('Section', '').strip() or
                            row.get('Section Des', '').strip() or
                            row.get('Section Description', '').strip()
                        )
                        # If still no section, try to use a default or skip
                        if not section_path:
                            # Try to use Suite or other fallback
                            section_path = row.get('Suite', '').strip()
                            if not section_path:
                                section_path = 'Imported'  # Default section name
                        
                        parent_section = importer.get_or_create_section_by_path(section_path)
                        
                        if not parent_section:
                            error_count += 1
                            errors.append(f"Row {row_num}: Could not create section")
                            continue
                        
                        # Create test case - skip empty rows
                        title = row.get('Title', '').strip()
                        if not title:
                            # Skip empty rows (no title means it's likely an empty row or header)
                            continue
                        
                        # Get field values with fallbacks for different column names
                        type_value = row.get('Type', '').strip() or 'Other'
                        priority_value = row.get('Priority', '').strip() or 'Medium'
                        estimate_value = row.get('Estimate', '').strip()
                        automation_type_value = (
                            row.get('Automation', '').strip() or
                            row.get('Automation Type', '').strip() or
                            'None'
                        )
                        labels_value = row.get('Labels', '').strip()
                        description_value = (
                            row.get('Description', '').strip() or
                            row.get('Mission', '').strip()
                        )
                        preconditions_value = (
                            row.get('Preconditions', '').strip() or
                            row.get('Precondition', '').strip()
                        )
                        
                        test_case = TestCase(
                            project=project,
                            title=title,
                            section=parent_section,
                            type=type_value or 'Other',
                            priority=priority_value or 'Medium',
                            estimate=estimate_value,
                            automation_type=automation_type_value or 'None',
                            labels=labels_value,
                            description=description_value,
                            preconditions=preconditions_value,
                            template='Test Case (Steps)',
                            is_deleted=False
                        )
                        
                        # Parse test steps - handle both formats
                        steps = []
                        
                        # Try TestRail format first - check "Steps (Addi)" which contains multiple steps
                        steps_column = (
                            row.get('Steps (Addi)', '').strip() or
                            row.get('Steps (Additional)', '').strip() or
                            row.get('Steps', '').strip()
                        )
                        if steps_column:
                            # Parse TestRail Steps format - "Steps (Addi)" contains multiple "Step Description: ... Expected Result: ..." entries
                            # Primary pattern: "Step Description:" followed by text, then "Expected Result:" or similar
                            step_desc_pattern = re.compile(
                                r'Step\s*Description[:.]?\s*(.*?)(?=Step\s*Description[:.]|Expected\s*(?:Result)?[:.]|$)', 
                                re.IGNORECASE | re.DOTALL
                            )
                            expected_pattern = re.compile(
                                r'Expected\s*(?:Result)?[:.]?\s*(.*?)(?=Step\s*Description[:.]|Expected\s*(?:Result)?[:.]|$)', 
                                re.IGNORECASE | re.DOTALL
                            )
                            
                            # Find all step descriptions and expected results
                            step_desc_matches = list(step_desc_pattern.finditer(steps_column))
                            expected_matches = list(expected_pattern.finditer(steps_column))
                            
                            if step_desc_matches:
                                # Match step descriptions with their expected results
                                for i, step_match in enumerate(step_desc_matches):
                                    step_desc = step_match.group(1).strip()
                                    expected_result = ''
                                    
                                    # Find the expected result that comes after this step description
                                    step_end = step_match.end()
                                    for exp_match in expected_matches:
                                        if exp_match.start() >= step_end:
                                            expected_result = exp_match.group(1).strip()
                                            break
                                    
                                    # If no expected result found after, try to find one before the next step
                                    if not expected_result:
                                        next_step_start = step_desc_matches[i + 1].start() if i + 1 < len(step_desc_matches) else len(steps_column)
                                        for exp_match in expected_matches:
                                            if step_end <= exp_match.start() < next_step_start:
                                                expected_result = exp_match.group(1).strip()
                                                break
                                    
                                    # Clean up step description (remove extra whitespace, newlines)
                                    step_desc = re.sub(r'\s+', ' ', step_desc).strip()
                                    expected_result = re.sub(r'\s+', ' ', expected_result).strip() if expected_result else ''
                                    
                                    # Remove common prefixes/suffixes that might be in the text
                                    step_desc = re.sub(r'^(The|Prepare|Capture|Validate)\s+', '', step_desc, flags=re.IGNORECASE).strip()
                                    
                                    if step_desc and len(step_desc) > 3:  # Minimum meaningful length
                                        steps.append(TestStep(
                                            step_number=len(steps) + 1,
                                            description=step_desc[:500],
                                            expected_result=expected_result[:500] if expected_result else 'See description'
                                        ))
                            
                            # If no "Step Description:" pattern found, try numbered steps
                            if not steps:
                                step_pattern = re.compile(r'Step\s*(\d+)[:.]?\s*(.*?)(?=Step\s*\d+[:.]|Expected|Step\s*Description|$)', re.IGNORECASE | re.DOTALL)
                                step_matches = list(step_pattern.finditer(steps_column))
                                
                                if step_matches:
                                    for match in step_matches:
                                        step_desc = match.group(2).strip()
                                        # Try to extract expected result from the step description
                                        expected_result = ''
                                        if 'Expected' in step_desc or 'expected' in step_desc:
                                            parts = re.split(r'[Ee]xpected\s*(?:Result)?[:.]?\s*', step_desc, 1)
                                            if len(parts) == 2:
                                                step_desc = parts[0].strip()
                                                expected_result = parts[1].strip()
                                        
                                        step_desc = re.sub(r'\s+', ' ', step_desc).strip()
                                        if step_desc and len(step_desc) > 3:
                                            steps.append(TestStep(
                                                step_number=len(steps) + 1,
                                                description=step_desc[:500],
                                                expected_result=expected_result[:500] if expected_result else 'See description'
                                            ))
                            
                            # If still no steps, try splitting by delimiters (##, ---, newlines)
                            if not steps and steps_column:
                                step_sections = re.split(r'\n\s*##\s*|\n\s*---\s*|\n\n+|\n\s*Request\s+Body', steps_column, flags=re.IGNORECASE)
                                for section in step_sections:
                                    section = section.strip()
                                    if section and len(section) > 10:
                                        if 'Expected' in section or 'expected' in section:
                                            parts = re.split(r'[Ee]xpected\s*(?:Result)?[:.]?\s*', section, 1)
                                            step_desc = parts[0].strip() if parts else section
                                            expected_result = parts[1].strip() if len(parts) > 1 else 'See description'
                                        else:
                                            step_desc = section
                                            expected_result = 'See description'
                                        
                                        step_desc = re.sub(r'\s+', ' ', step_desc).strip()
                                        if step_desc and len(step_desc) > 5:
                                            steps.append(TestStep(
                                                step_number=len(steps) + 1,
                                                description=step_desc[:500],
                                                expected_result=expected_result[:500]
                                            ))
                        
                        # If no steps from Steps column, try separate step columns
                        if not steps:
                            for i in range(1, 20):
                                step_desc = row.get(f'Step {i}', '').strip()
                                if not step_desc:
                                    step_desc = row.get(f'Steps (Step) {i}', '').strip()
                                if not step_desc:
                                    continue
                                
                                expected = row.get(f'Expected Result {i}', '').strip()
                                if not expected:
                                    expected = row.get(f'Expected {i}', '').strip()
                                
                                if step_desc:
                                    steps.append(TestStep(
                                        step_number=len(steps) + 1,
                                        description=step_desc,
                                        expected_result=expected or 'See description'
                                    ))
                        
                        # If still no steps, try other step-related columns
                        if not steps:
                            # Try "Steps (Step)" column
                            steps_step_col = row.get('Steps (Step)', '').strip()
                            if steps_step_col:
                                steps.append(TestStep(
                                    step_number=1,
                                    description=steps_step_col[:500],
                                    expected_result='See description'
                                ))
                            
                            # Try "Steps (Expe)" column for expected results
                            steps_expe_col = row.get('Steps (Expe)', '').strip()
                            if steps_expe_col and steps:
                                steps[0].expected_result = steps_expe_col[:500]
                        
                        # If still no steps, try to create one from any Steps column as-is
                        if not steps and steps_column:
                            # Use the entire Steps column as one step
                            steps.append(TestStep(
                                step_number=1,
                                description=steps_column[:500],  # Limit length
                                expected_result='See description'
                            ))
                        
                        # If still no steps, try to use other columns that might contain step info
                        if not steps:
                            # Try "Precondition" as a step if it exists
                            preconditions = row.get('Precondition', '').strip() or row.get('Preconditions', '').strip()
                            if preconditions and len(preconditions) > 10:
                                steps.append(TestStep(
                                    step_number=1,
                                    description=f"Precondition: {preconditions[:450]}",
                                    expected_result='See description'
                                ))
                        
                        if not steps:
                            error_count += 1
                            errors.append(f"Row {row_num}: At least one test step is required")
                            continue
                        
                        test_case.steps = steps
                        
                        if importer.queue_test_case(test_case):
                            imported_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
            
            elif file_name.endswith(('.xlsx', '.xls')):
                # Process Excel file
                from openpyxl import load_workbook
                
                wb = load_workbook(uploaded_file)
                ws = wb.active
                
                # Read header row
                headers = [cell.value for cell in ws[1]]
                header_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header}
                
                # First pass: Collect all rows and group by ID (to handle multi-row test cases)
                # Continuation rows may not have ID, so we track the last seen ID
                rows_by_id = {}
                current_test_id = None
                current_test_rows = []
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    # Convert row to dict
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        if header and col_idx < len(row):
                            row_dict[str(header).strip()] = row[col_idx].value if row[col_idx].value else ''
                    
                    # Get ID and Title
                    test_id = str(row_dict.get('ID', '')).strip()
                    title = str(row_dict.get('Title', '')).strip()
                    
                    # If this row has an ID, it's either a new test case or continuation
                    if test_id:
                        # If we have a previous test case being built, save it
                        if current_test_id and current_test_rows:
                            rows_by_id[current_test_id] = current_test_rows
                        
                        # Start a new test case group
                        current_test_id = test_id
                        current_test_rows = [(row_num, row_dict)]
                    elif title:
                        # Row has title but no ID - this is a new test case (use title as ID for grouping)
                        # If we have a previous test case being built, save it
                        if current_test_id and current_test_rows:
                            rows_by_id[current_test_id] = current_test_rows
                        
                        # Use title as temporary ID for grouping
                        current_test_id = f"TEMP_{title[:50]}"  # Use title as grouping key
                        current_test_rows = [(row_num, row_dict)]
                    else:
                        # This is a continuation row (no ID, no title) - add to current test case
                        if current_test_id and current_test_rows:
                            current_test_rows.append((row_num, row_dict))
                        # If no current test case, skip this row (orphaned continuation row)
                
                # Don't forget the last test case
                if current_test_id and current_test_rows:
                    rows_by_id[current_test_id] = current_test_rows
                
                # Second pass: Process each test case (grouped by ID)
                # IMPORTANT: Only create ONE test case per ID group, collecting steps from ALL rows
                for test_id, id_rows in rows_by_id.items():
                    importer.stats['rows_processed'] += 1
                    try:
                        # Find the main row (first row with a title AND other metadata)
                        # This is the row that defines the test case
                        main_row = None
                        main_row_num = None
                        for row_num, row_dict in id_rows:
                            title = str(row_dict.get('Title', '')).strip()
                            # Check if this row has both title AND other metadata (not just a continuation row with accidental title)
                            has_metadata = (
                                title and (
                                    str(row_dict.get('Precondition', '')).strip() or
                                    str(row_dict.get('Priority', '')).strip() or
                                    str(row_dict.get('Section', '')).strip() or
                                    str(row_dict.get('Section Hie', '')).strip()
                                )
                            )
                            if has_metadata:
                                main_row = row_dict
                                main_row_num = row_num
                                break
                        
                        # If still no main row, try just finding first row with title
                        if not main_row:
                            for row_num, row_dict in id_rows:
                                title = str(row_dict.get('Title', '')).strip()
                                if title:
                                    main_row = row_dict
                                    main_row_num = row_num
                                    break
                        
                        # If no main row found, skip this ID group
                        if not main_row:
                            continue
                        
                        # Use main_row for test case data
                        row_dict = main_row
                        row_num = main_row_num
                        
                        # Get or create section - try multiple column name variations
                        # Priority: Section Hie (Section Hierarchy) > Section > Section Des > Suite
                        section_path = (
                            str(row_dict.get('Section Hie', '')).strip() or
                            str(row_dict.get('Section Hier', '')).strip() or
                            str(row_dict.get('Section Hierarchical', '')).strip() or
                            str(row_dict.get('Section', '')).strip() or
                            str(row_dict.get('Section Des', '')).strip() or
                            str(row_dict.get('Section Description', '')).strip()
                        )
                        # If still no section, try to use a default or skip
                        if not section_path:
                            # Try to use Suite or other fallback
                            section_path = str(row_dict.get('Suite', '')).strip()
                            if not section_path:
                                section_path = 'Imported'  # Default section name
                        
                        parent_section = importer.get_or_create_section_by_path(section_path)
                        
                        if not parent_section:
                            error_count += 1
                            errors.append(f"Row {row_num}: Could not create section")
                            continue
                        
                        # Create test case - skip empty rows
                        title = str(row_dict.get('Title', '')).strip()
                        if not title:
                            # Skip empty rows (no title means it's likely an empty row or header)
                            continue
                        
                        # Get field values with fallbacks for different column names
                        type_value = (
                            str(row_dict.get('Type', '')).strip() or
                            'Other'
                        )
                        priority_value = (
                            str(row_dict.get('Priority', '')).strip() or
                            'Medium'
                        )
                        estimate_value = str(row_dict.get('Estimate', '')).strip()
                        automation_type_value = (
                            str(row_dict.get('Automation', '')).strip() or
                            str(row_dict.get('Automation Type', '')).strip() or
                            'None'
                        )
                        labels_value = str(row_dict.get('Labels', '')).strip()
                        description_value = (
                            str(row_dict.get('Description', '')).strip() or
                            str(row_dict.get('Mission', '')).strip()
                        )
                        preconditions_value = (
                            str(row_dict.get('Preconditions', '')).strip() or
                            str(row_dict.get('Precondition', '')).strip()
                        )
                        
                        test_case = TestCase(
                            project=project,
                            title=title,
                            section=parent_section,
                            type=type_value or 'Other',
                            priority=priority_value or 'Medium',
                            estimate=estimate_value,
                            automation_type=automation_type_value or 'None',
                            labels=labels_value,
                            description=description_value,
                            preconditions=preconditions_value,
                            template='Test Case (Steps)',
                            is_deleted=False
                        )
                        
                        # Parse test steps - collect from ALL rows with the same ID
                        steps = []
                        
                        # Process all rows for this test case ID to collect steps
                        for step_row_num, step_row_dict in id_rows:
                            # Check if this row has step data - look for any step-related content
                            has_step_data = False
                            
                            # Try TestRail format - check all step columns from this row
                            steps_columns_to_check = [
                                ('Steps (Addi)', 'Steps (Additional)'),
                                ('Steps (Expe)', 'Steps (Expected)'),
                                ('Steps (Refer)', 'Steps (Reference)'),
                                ('Steps (Shar)', 'Steps (Shared)'),
                                ('Steps (Step)', 'Steps (Step Description)'),
                                ('Steps',),
                            ]
                            
                            # Also check for columns that might contain step data (based on screenshot)
                            # Column U might be "Steps" and Column V might be "Expected Result"
                            # Try various column name variations (case-insensitive)
                            step_col = ''
                            expected_col = ''
                            
                            # Try to find Steps column with various name variations
                            for col_name, col_value in step_row_dict.items():
                                col_name_lower = col_name.lower().strip()
                                if col_name_lower == 'steps' or col_name_lower.startswith('steps ('):
                                    step_col = str(col_value).strip()
                                    break
                            
                            # If not found, try exact match
                            if not step_col:
                                step_col = str(step_row_dict.get('Steps', '')).strip()
                            
                            # Try to find Expected Result column with various name variations
                            for col_name, col_value in step_row_dict.items():
                                col_name_lower = col_name.lower().strip()
                                if 'expected' in col_name_lower and 'result' in col_name_lower:
                                    expected_col = str(col_value).strip()
                                    break
                            
                            # If not found, try exact matches
                            if not expected_col:
                                expected_col = str(step_row_dict.get('Expected Result', '')).strip() or str(step_row_dict.get('Expected R.', '')).strip()
                            
                            # If we have step or expected result data, process it
                            if step_col or expected_col:
                                # Check if step_col contains both description and expected result
                                # Pattern: "Step Description: ... Expected Result: ..."
                                # First check if "Expected Result:" appears in step_col
                                if step_col and re.search(r'Expected\s*(?:Result)?[:.]', step_col, re.IGNORECASE):
                                    # Step column contains both description and expected result
                                    # Split by "Expected Result:" pattern - everything before goes to description, everything after goes to expected_result
                                    # Handle cases with or without "Result", with various spacing/newlines
                                    parts = re.split(r'Expected\s+(?:Result\s*)?[:.]\s*', step_col, 1, flags=re.IGNORECASE | re.DOTALL)
                                    if len(parts) == 2:
                                        # Extract description (everything before "Expected Result:")
                                        step_desc = parts[0].strip()
                                        # Remove "Step Description:" prefix if present
                                        step_desc = re.sub(r'^Step\s*Description[:.]?\s*', '', step_desc, flags=re.IGNORECASE).strip()
                                        
                                        # Extract expected result (everything after "Expected Result:")
                                        expected_result = parts[1].strip()
                                        
                                        # Clean up whitespace (replace multiple spaces/newlines with single space)
                                        step_desc = re.sub(r'\s+', ' ', step_desc).strip()
                                        expected_result = re.sub(r'\s+', ' ', expected_result).strip()
                                        
                                        # Only use expected_col if expected_result from step_col is empty
                                        if not expected_result and expected_col:
                                            expected_result = expected_col
                                            expected_result = re.sub(r'^Expected\s*(?:Result)?[:.]?\s*', '', expected_result, flags=re.IGNORECASE).strip()
                                            expected_result = re.sub(r'\s+', ' ', expected_result).strip()
                                        
                                        if step_desc and len(step_desc.strip()) > 0:
                                            steps.append(TestStep(
                                                step_number=len(steps) + 1,
                                                description=step_desc[:500],
                                                expected_result=expected_result[:500] if expected_result else 'See description'
                                            ))
                                            has_step_data = True
                                elif step_col:
                                    # Step column only contains description (no "Expected Result:" in it)
                                    # Use expected_col for expected result
                                    step_desc = step_col
                                    # Remove "Step Description:" prefix if present
                                    step_desc = re.sub(r'^Step\s*Description[:.]?\s*', '', step_desc, flags=re.IGNORECASE).strip()
                                    step_desc = re.sub(r'\s+', ' ', step_desc).strip()
                                    
                                    if step_desc and len(step_desc.strip()) > 0:
                                        expected_result = expected_col
                                        # Remove "Expected Result:" prefix if present
                                        if expected_result:
                                            expected_result = re.sub(r'^Expected\s*(?:Result)?[:.]?\s*', '', expected_result, flags=re.IGNORECASE).strip()
                                            expected_result = re.sub(r'\s+', ' ', expected_result).strip()
                                        
                                        steps.append(TestStep(
                                            step_number=len(steps) + 1,
                                            description=step_desc[:500],
                                            expected_result=expected_result[:500] if expected_result else 'See description'
                                        ))
                                        has_step_data = True
                            
                            # Also check the other step columns
                            for col_variations in steps_columns_to_check:
                                steps_column = ''
                                for col_name in col_variations:
                                    steps_column = str(step_row_dict.get(col_name, '')).strip()
                                    if steps_column:
                                        break
                                
                                if steps_column and not has_step_data:
                                    # Parse "Step Description:" pattern
                                    step_desc_pattern = re.compile(
                                        r'Step\s*Description[:.]?\s*(.*?)(?=Step\s*Description[:.]|Expected\s*(?:Result)?[:.]|$)', 
                                        re.IGNORECASE | re.DOTALL
                                    )
                                    expected_pattern = re.compile(
                                        r'Expected\s*(?:Result)?[:.]?\s*(.*?)(?=Step\s*Description[:.]|Expected\s*(?:Result)?[:.]|$)', 
                                        re.IGNORECASE | re.DOTALL
                                    )
                                    
                                    step_desc_matches = list(step_desc_pattern.finditer(steps_column))
                                    expected_matches = list(expected_pattern.finditer(steps_column))
                                    
                                    if step_desc_matches:
                                        for i, step_match in enumerate(step_desc_matches):
                                            step_desc = step_match.group(1).strip()
                                            expected_result = ''
                                            
                                            step_end = step_match.end()
                                            for exp_match in expected_matches:
                                                if exp_match.start() >= step_end:
                                                    expected_result = exp_match.group(1).strip()
                                                    break
                                            
                                            if not expected_result:
                                                next_step_start = step_desc_matches[i + 1].start() if i + 1 < len(step_desc_matches) else len(steps_column)
                                                for exp_match in expected_matches:
                                                    if step_end <= exp_match.start() < next_step_start:
                                                        expected_result = exp_match.group(1).strip()
                                                        break
                                            
                                            step_desc = re.sub(r'\s+', ' ', step_desc).strip()
                                            expected_result = re.sub(r'\s+', ' ', expected_result).strip() if expected_result else ''
                                            
                                            if step_desc and len(step_desc.strip()) > 0:
                                                steps.append(TestStep(
                                                    step_number=len(steps) + 1,
                                                    description=step_desc[:500],
                                                    expected_result=expected_result[:500] if expected_result else 'See description'
                                                ))
                                    else:
                                        # If no "Step Description:" pattern, use the entire column as one step
                                        if steps_column and len(steps_column.strip()) > 0:
                                            steps.append(TestStep(
                                                step_number=len(steps) + 1,
                                                description=steps_column[:500],
                                                expected_result='See description'
                                            ))
                                    break  # Found steps in this column, move to next row
                        
                        # If still no steps, try separate step columns from main row
                        if not steps:
                            # Try various column name patterns
                            for i in range(1, 20):
                                step_desc = str(row_dict.get(f'Step {i}', '')).strip()
                                if not step_desc:
                                    step_desc = str(row_dict.get(f'Steps (Step) {i}', '')).strip()
                                if not step_desc:
                                    step_desc = str(row_dict.get(f'Step Description {i}', '')).strip()
                                if not step_desc:
                                    # Try case-insensitive search in all columns
                                    for col_name, col_value in row_dict.items():
                                        if col_name.lower() == f'step {i}'.lower() or col_name.lower() == f'steps (step) {i}'.lower():
                                            step_desc = str(col_value).strip()
                                            break
                                if not step_desc:
                                    continue
                                
                                expected = str(row_dict.get(f'Expected Result {i}', '')).strip()
                                if not expected:
                                    expected = str(row_dict.get(f'Expected {i}', '')).strip()
                                if not expected:
                                    # Try case-insensitive search
                                    for col_name, col_value in row_dict.items():
                                        if col_name.lower() == f'expected result {i}'.lower() or col_name.lower() == f'expected {i}'.lower():
                                            expected = str(col_value).strip()
                                            break
                                
                                if step_desc and len(step_desc) > 0:
                                    steps.append(TestStep(
                                        step_number=len(steps) + 1,
                                        description=step_desc[:500],
                                        expected_result=(expected[:500] if expected else 'See description')
                                    ))
                        
                        # Additional fallback: check if Description column can be used as a step
                        if not steps:
                            description_value = str(row_dict.get('Description', '')).strip()
                            if description_value and len(description_value) > 10:
                                steps.append(TestStep(
                                    step_number=1,
                                    description=description_value[:500],
                                    expected_result='See description'
                                ))
                        
                        # Final fallback: use precondition as step
                        if not steps:
                            preconditions = str(row_dict.get('Precondition', '')).strip() or str(row_dict.get('Preconditions', '')).strip()
                            if preconditions and len(preconditions) > 10:
                                steps.append(TestStep(
                                    step_number=1,
                                    description=f"Precondition: {preconditions[:450]}",
                                    expected_result='See description'
                                ))
                        
                        # Last resort: create a minimal step from title if nothing else works
                        if not steps:
                            title_value = str(row_dict.get('Title', '')).strip()
                            if title_value and len(title_value) > 5:
                                steps.append(TestStep(
                                    step_number=1,
                                    description=f"Execute test case: {title_value[:450]}",
                                    expected_result='See description'
                                ))
                        
                        if not steps:
                            error_count += 1
                            errors.append(f"Row {row_num}: At least one test step is required")
                            continue
                        
                        test_case.steps = steps
                        
                        if importer.queue_test_case(test_case):
                            imported_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
            
            else:
                # Attempt XML import
                try:
                    uploaded_file.seek(0)
                except Exception:
                    pass
                try:
                    xml_root = ET.parse(uploaded_file).getroot()
                except Exception as xml_error:
                    messages.error(request, f'Unsupported file format or XML parse error: {xml_error}')
                return redirect('testcases:test_case_list', project_id=str(project.id))
            
                def clean_text(value):
                    return html.unescape((value or '').strip())
                
                def parse_xml_steps(case_elem):
                    steps = []
                    steps_container = case_elem.find('.//steps_separated')
                    if steps_container is not None:
                        for idx, step_elem in enumerate(steps_container.findall('step'), start=1):
                            desc = clean_text(step_elem.findtext('content', ''))
                            expected = clean_text(step_elem.findtext('expected', ''))
                            if desc:
                                steps.append(TestStep(
                                    step_number=idx,
                                    description=desc[:500],
                                    expected_result=(expected or 'See description')[:500]
                                ))
                    return steps
                
                def process_section(section_elem, path_parts):
                    nonlocal imported_count, error_count

                    section_name = clean_text(section_elem.findtext('name', ''))
                    if not section_name:
                        return
                    current_path = '>'.join(path_parts + [section_name])
                    parent_section = importer.get_or_create_section_by_path(current_path)
                    
                    cases_elem = section_elem.find('cases')
                    if cases_elem is not None:
                        for case_elem in cases_elem.findall('case'):
                            importer.stats['rows_processed'] += 1
                            title = clean_text(case_elem.findtext('title', ''))
                            if not title:
                                continue
                            description_value = clean_text(case_elem.findtext('description', ''))
                            type_value = clean_text(case_elem.findtext('type', '')) or 'Other'
                            priority_value = clean_text(case_elem.findtext('priority', '')) or 'Medium'
                            estimate_value = clean_text(case_elem.findtext('estimate', ''))
                            automation_type_value = clean_text(case_elem.findtext('.//automation_type/value', '')) or 'None'
                            preconditions_value = clean_text(case_elem.findtext('.//preconds', ''))
                            
                            steps = parse_xml_steps(case_elem)
                            if not steps:
                                if description_value:
                                    steps = [TestStep(step_number=1, description=description_value[:500], expected_result='See description')]
                                else:
                                    error_count += 1
                                    errors.append(f"XML Case '{title}': Missing steps")
                                    continue
                            
                            test_case = TestCase(
                                project=project,
                                title=title,
                                section=parent_section,
                                type=type_value,
                                priority=priority_value,
                                estimate=estimate_value,
                                automation_type=automation_type_value,
                                labels='',
                                description=description_value,
                                preconditions=preconditions_value,
                                template='Test Case (Steps)',
                                is_deleted=False,
                                steps=steps
                            )
                            
                            if importer.queue_test_case(test_case):
                                imported_count += 1
                    
                    subsections = section_elem.find('sections')
                    if subsections is not None:
                        for child in subsections.findall('section'):
                            process_section(child, path_parts + [section_name])
                
                suite_sections = xml_root.find('sections')
                if suite_sections is not None:
                    for top_section in suite_sections.findall('section'):
                        process_section(top_section, [])
                else:
                    errors.append('XML file missing <sections> root.')

            stats = importer.finalize()
            imported_count = stats['cases_created']
            total_rows = stats['rows_processed']
            duplicates = stats['duplicates_skipped']
            new_sections = stats['sections_created']
            total_steps = stats['steps_created']
            
            if imported_count > 0:
                info_parts = [
                    f'{imported_count} test case(s)',
                    f'{total_rows} row(s) processed',
                    f'{total_steps} step(s) captured'
                ]
                if new_sections:
                    info_parts.append(f'{new_sections} section(s) created')
                if duplicates:
                    info_parts.append(f'{duplicates} duplicate(s) skipped')
                messages.success(request, 'Import complete: ' + ', '.join(info_parts) + '.')
            combined_errors = errors + stats['errors']
            if error_count or combined_errors:
                total_issues = error_count + len(stats['errors'])
                preview = combined_errors[:5]
                messages.warning(request, f'Encountered {total_issues} issue(s) during import: {", ".join(preview)}')
            
        except Exception as e:
            messages.error(request, f'Error importing file: {str(e)}')
        
        return redirect('testcases:test_case_list', project_id=str(project.id))
    
    # GET request - show import form
    context = {
        'project': project,
    }
    return render(request, 'testcases/test_case_import.html', context)


@login_required
def project_dashboard(request):
    """Project dashboard - shows all projects (similar to TestRail dashboard)"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Get all active projects
    projects = Project.objects(is_active=True).order_by('name')
    
    # Get statistics for each project
    projects_data = []
    for project in projects:
        projects_data.append({
            'project': project,
            'test_case_count': project.get_test_case_count(),
            'section_count': project.get_section_count(),
            'active_test_runs': project.get_active_test_runs_count(),
            'active_milestones': project.get_active_milestones_count(),
        })
    
    context = {
        'projects': projects_data,
        'total_projects': len(projects_data),
    }
    return render(request, 'testcases/project_dashboard.html', context)


@login_required
def project_list(request):
    """List all projects"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    projects = Project.objects().order_by('name')
    
    context = {
        'projects': projects,
    }
    return render(request, 'testcases/project_list.html', context)


@login_required
def project_create(request):
    """Create a new project - Admin only"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Only admins can create projects
    if request.user.role != 3:
        messages.error(request, 'Only administrators can create projects.')
        return redirect('testcases:project_dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        
        if name:
            try:
                # Check if project with same name exists
                existing = Project.objects(name=name).first()
                if existing:
                    messages.error(request, f'Project "{name}" already exists.')
                else:
                    project = Project(
                        name=name,
                        description=description,
                        is_active=True,
                        created_by_id=request.user.id
                    )
                    project.save()
                    
                    # Record change history
                    record_change('project', str(project.id), 'created', request.user.id, f'Created project: {project.name}')
                    
                    messages.success(request, f'Project "{name}" created successfully!')
                    return redirect('testcases:project_dashboard')
            except Exception as e:
                messages.error(request, f'Error creating project: {str(e)}')
        else:
            messages.error(request, 'Project name is required.')
    
    return render(request, 'testcases/project_create.html')


@login_required
def project_edit(request, project_id):
    """Edit a project - Admin only"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Only admins can edit projects
    if request.user.role != 3:
        messages.error(request, 'Only administrators can edit projects.')
        return redirect('testcases:project_dashboard')
    
    try:
        project = Project.objects.get(id=ObjectId(project_id))
    except (Project.DoesNotExist, Exception):
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        if name:
            # Check if another project with same name exists
            existing = Project.objects(name=name).first()
            if existing and str(existing.id) != str(project.id):
                messages.error(request, f'Project "{name}" already exists.')
            else:
                project.name = name
                project.description = description
                project.is_active = is_active
                project.updated_by_id = request.user.id
                project.save()
                
                # Record change history
                record_change('project', str(project.id), 'updated', request.user.id, f'Updated project: {project.name}')
                
                messages.success(request, f'Project "{project.name}" updated successfully!')
                return redirect('testcases:project_dashboard')
        else:
            messages.error(request, 'Project name is required.')
    
    context = {
        'project': project,
    }
    return render(request, 'testcases/project_edit.html', context)


@login_required
@require_http_methods(["POST"])
def project_delete(request, project_id):
    """Delete a project (soft delete by setting is_active=False)"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    # Only admins can delete projects
    if request.user.role != 3:
        messages.error(request, 'Only administrators can delete projects.')
        return redirect('testcases:project_dashboard')
    
    try:
        project = Project.objects.get(id=ObjectId(project_id))
        project_name = project.name
        project.is_active = False
        project.updated_by_id = request.user.id
        project.save()
        
        # Record change history
        record_change('project', str(project.id), 'deleted', request.user.id, f'Deleted project: {project_name}')
        
        messages.success(request, f'Project "{project_name}" deleted successfully!')
    except (Project.DoesNotExist, Exception):
        messages.error(request, 'Project not found.')
    
    return redirect('testcases:project_dashboard')


# ==================== Test Run Views ====================

@login_required
def test_run_list(request, project_id):
    """List all test runs for a project"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Get filter parameters
    filter_status = request.GET.get('filter', 'open')  # 'open', 'closed', or 'all'
    
    # Get test runs based on filter
    if filter_status == 'open':
        test_runs = TestRun.objects(project=project, is_closed=False).order_by('-created_at')
    elif filter_status == 'closed':
        test_runs = TestRun.objects(project=project, is_closed=True).order_by('-created_at')
    else:
        test_runs = TestRun.objects(project=project).order_by('-created_at')
    
    # Calculate statistics for each test run
    test_runs_data = []
    for test_run in test_runs:
        summary = test_run.get_results_summary()
        test_runs_data.append({
            'test_run': test_run,
            'summary': summary,
        })
    
    # Count open and closed test runs
    open_count = TestRun.objects(project=project, is_closed=False).count()
    closed_count = TestRun.objects(project=project, is_closed=True).count()
    
    context = {
        'project': project,
        'test_runs_data': test_runs_data,
        'open_count': open_count,
        'closed_count': closed_count,
        'filter_status': filter_status,
    }
    return render(request, 'testcases/test_run_list.html', context)


@login_required
def test_run_add(request, project_id):
    """Add a new test run"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        references = request.POST.get('references', '').strip()
        milestone_id = request.POST.get('milestone', '').strip()
        assigned_to_id = request.POST.get('assigned_to', '').strip()
        start_date_str = request.POST.get('start_date', '').strip()
        end_date_str = request.POST.get('end_date', '').strip()
        inclusion_type = request.POST.get('inclusion_type', 'all').strip()
        
        # Get test case IDs - handle both array and comma-separated string
        test_case_ids = []
        raw_id_sources = []
        test_case_ids_input = request.POST.get('test_case_ids', '').strip()
        if test_case_ids_input:
            raw_id_sources.append(test_case_ids_input)
        raw_id_sources.extend(filter(None, request.POST.getlist('test_case_ids[]')))
        for source in raw_id_sources:
            for tc_id in source.split(','):
                tc_id = tc_id.strip()
                if tc_id:
                    test_case_ids.append(tc_id)
        
        # Validate required fields
        if not name:
            messages.error(request, 'Test run name is required.')
            from django.contrib.auth import get_user_model
            User = get_user_model()
            users = User.objects.filter(is_active=True).order_by('username')
            return render(request, 'testcases/test_run_add.html', {'project': project, 'users': users})
        
        # Parse dates
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                pass
        
        # Parse assigned_to_id
        assigned_to = None
        if assigned_to_id:
            try:
                assigned_to = int(assigned_to_id)
            except ValueError:
                pass
        
        # Create test run
        test_run = TestRun(
            project=project,
            name=name,
            description=description,
            references=references,
            milestone_id=milestone_id,
            assigned_to_id=assigned_to,
            start_date=start_date,
            end_date=end_date,
            inclusion_type=inclusion_type,
            test_case_ids=test_case_ids if inclusion_type == 'specific' else [],
            created_by_id=request.user.id,
            is_closed=False
        )
        test_run.save()
        
        # Log test run creation
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[TEST_RUN_ADD] Created test run: ID={test_run.id}, Name={test_run.name}, Project={project.id}, Inclusion Type={inclusion_type}")
        print(f"[TEST_RUN_ADD] Created test run: ID={test_run.id}, Name={test_run.name}, Project={project.id}, Inclusion Type={inclusion_type}")
        
        # Check if project has any test cases at all
        from .models import TestCase
        total_project_test_cases = TestCase.objects(project=project, is_deleted=False).count()
        logger.info(f"[TEST_RUN_ADD] Total test cases in project: {total_project_test_cases}")
        print(f"[TEST_RUN_ADD] Total test cases in project: {total_project_test_cases}")
        
        # Update results for test cases - this creates TestRunResult entries for each test case
        try:
            # First, verify we can get test cases
            logger.info(f"[TEST_RUN_ADD] Calling get_test_cases() for test run {test_run.id}")
            print(f"[TEST_RUN_ADD] Calling get_test_cases() for test run {test_run.id}")
            test_cases = test_run.get_test_cases()
            
            # Convert to list to count properly
            if hasattr(test_cases, '__iter__') and not isinstance(test_cases, (list, tuple)):
                test_cases_list = list(test_cases)
            else:
                test_cases_list = test_cases if isinstance(test_cases, list) else list(test_cases)
            
            test_case_count = len(test_cases_list)
            logger.info(f"[TEST_RUN_ADD] Found {test_case_count} test cases for test run {test_run.id}")
            print(f"[TEST_RUN_ADD] Found {test_case_count} test cases for test run {test_run.id}")
            
            if test_case_count == 0:
                if inclusion_type == 'all':
                    logger.warning(f"[TEST_RUN_ADD] No test cases found in project {project.id} for 'all' inclusion type")
                    print(f"[TEST_RUN_ADD] WARNING: No test cases found in project {project.id} for 'all' inclusion type")
                    messages.warning(request, 'Test run created, but no test cases found in this project. Please add test cases first.')
                elif inclusion_type == 'specific':
                    logger.warning(f"[TEST_RUN_ADD] No test cases selected for 'specific' inclusion type. test_case_ids: {test_run.test_case_ids}")
                    print(f"[TEST_RUN_ADD] WARNING: No test cases selected. test_case_ids: {test_run.test_case_ids}")
                    messages.warning(request, 'Test run created, but no test cases were selected. Please select test cases when creating a test run.')
                else:
                    logger.warning(f"[TEST_RUN_ADD] No test cases match filter criteria for 'dynamic' inclusion type")
                    print(f"[TEST_RUN_ADD] WARNING: No test cases match filter criteria")
                    messages.warning(request, 'Test run created, but no test cases match the filter criteria.')
            else:
                # Update results for test cases
                logger.info(f"[TEST_RUN_ADD] Calling update_results_for_test_cases() for {test_case_count} test cases")
                print(f"[TEST_RUN_ADD] Calling update_results_for_test_cases() for {test_case_count} test cases")
                results_count = test_run.update_results_for_test_cases()
                logger.info(f"[TEST_RUN_ADD] Created {results_count} results for test run {test_run.id}")
                print(f"[TEST_RUN_ADD] Created {results_count} results for test run {test_run.id}")
                
                # Verify results were saved
                test_run.reload()
                logger.info(f"[TEST_RUN_ADD] After update, test run has {len(test_run.results)} results")
                print(f"[TEST_RUN_ADD] After update, test run has {len(test_run.results)} results")
                
                messages.success(request, f'Successfully added the new test run with {test_case_count} test case(s).')
        except Exception as e:
            # Log error and show to user
            import traceback
            error_trace = traceback.format_exc()
            logger.error(f"[TEST_RUN_ADD] ERROR updating test case results for test run {test_run.id}: {str(e)}\n{error_trace}")
            print(f"[TEST_RUN_ADD] ERROR: {str(e)}")
            print(f"[TEST_RUN_ADD] TRACEBACK:\n{error_trace}")
            messages.warning(request, f'Test run created, but there was an issue connecting test cases: {str(e)}. Please check the test run and try again.')
        
        # Record change history
        record_change('test_run', str(test_run.id), 'created', request.user.id, f'Created test run: {test_run.name}')
        
        if 'Successfully added' not in [msg.message for msg in messages.get_messages(request)]:
            messages.success(request, 'Successfully added the new test run.')
        
        return redirect('testcases:test_run_detail', project_id=str(project.id), test_run_id=str(test_run.id))
    
    # GET request - show form
    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.filter(is_active=True).order_by('username')
    
    context = {
        'project': project,
        'users': users,
    }
    return render(request, 'testcases/test_run_add.html', context)


@login_required
def test_run_detail(request, project_id, test_run_id):
    """View test run details and results"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Validate test_run_id
    if not test_run_id:
        messages.error(request, 'Test run ID is required.')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    
    # Convert to string and strip
    test_run_id_str = str(test_run_id).strip().lower()
    
    # Check for reserved words that shouldn't be used as test_run_id
    reserved_words = ['add', 'select-cases', 'edit', 'delete', 'update-result', 'new', 'create']
    if test_run_id_str in reserved_words:
        # This is likely a URL routing issue - redirect to the appropriate page
        if test_run_id_str == 'add':
            return redirect('testcases:test_run_add', project_id=str(project.id))
        elif test_run_id_str == 'select-cases':
            return redirect('testcases:test_run_select_cases', project_id=str(project.id))
        else:
            messages.error(request, f'Invalid test run ID: "{test_run_id}". This is a reserved word.')
            return redirect('testcases:test_run_list', project_id=str(project.id))
    
    # Restore original case for ObjectId validation
    test_run_id_str = str(test_run_id).strip()
    
    if not test_run_id_str:
        messages.error(request, 'Test run ID cannot be empty.')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    
    try:
        # Convert test_run_id to ObjectId - try conversion and catch errors
        try:
            test_run_object_id = ObjectId(test_run_id_str)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Invalid test run ID format: {test_run_id_str}, error: {str(e)}")
            messages.error(request, f'Invalid test run ID format: "{test_run_id_str}". Please check the URL and ensure you are accessing a valid test run.')
            return redirect('testcases:test_run_list', project_id=str(project.id))
        
        test_run = TestRun.objects.get(id=test_run_object_id, project=project)
    except TestRun.DoesNotExist:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Test run not found: ID={test_run_id_str}, Project={project.id}")
        messages.error(request, f'Test run not found with ID: {test_run_id_str}')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading test run {test_run_id_str}: {str(e)}")
        messages.error(request, f'Error loading test run: {str(e)}')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    
    # Get test cases and their results
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[TEST_RUN_DETAIL] Loading test run {test_run.id}, inclusion_type={test_run.inclusion_type}")
    print(f"[TEST_RUN_DETAIL] Loading test run {test_run.id}, inclusion_type={test_run.inclusion_type}")
    
    try:
        logger.info(f"[TEST_RUN_DETAIL] Calling get_test_cases()")
        print(f"[TEST_RUN_DETAIL] Calling get_test_cases()")
        test_cases = test_run.get_test_cases()
        # Convert to list if it's a queryset
        if hasattr(test_cases, '__iter__') and not isinstance(test_cases, (list, tuple)):
            test_cases = list(test_cases)
        logger.info(f"[TEST_RUN_DETAIL] Retrieved {len(test_cases)} test cases")
        print(f"[TEST_RUN_DETAIL] Retrieved {len(test_cases)} test cases")
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"[TEST_RUN_DETAIL] Error getting test cases for test run {test_run.id}: {str(e)}\n{error_trace}")
        print(f"[TEST_RUN_DETAIL] ERROR getting test cases: {str(e)}")
        messages.warning(request, f'Error loading test cases: {str(e)}')
        test_cases = []
    
    # Ensure the results list stays in sync with the current set of test cases.
    if (not test_run.is_closed) and test_cases:
        expected_ids = {str(tc.id) for tc in test_cases}
        result_ids = {str(r.test_case.id) for r in test_run.results}
        if expected_ids != result_ids:
            logger.info(f"[TEST_RUN_DETAIL] Results/test-case mismatch detected. Syncing results...")
            print(f"[TEST_RUN_DETAIL] Results/test-case mismatch detected. Syncing results...")
            try:
                test_run.update_results_for_test_cases()
                test_run.reload()
                synced_cases = test_run.get_test_cases()
                if hasattr(synced_cases, '__iter__') and not isinstance(synced_cases, (list, tuple)):
                    synced_cases = list(synced_cases)
                test_cases = synced_cases
                logger.info(f"[TEST_RUN_DETAIL] Sync complete. Results now: {len(test_run.results)}")
                print(f"[TEST_RUN_DETAIL] Sync complete. Results now: {len(test_run.results)}")
                messages.info(request, 'Newly imported or removed test cases were synced into this test run.')
            except Exception as sync_error:
                logger.error(f"[TEST_RUN_DETAIL] Error auto-syncing test cases for run {test_run.id}: {sync_error}")
                print(f"[TEST_RUN_DETAIL] ERROR auto-syncing: {sync_error}")
    
    # If still empty (e.g., brand-new run in empty project) attempt a refresh for 'all' inclusion runs
    if not test_cases and test_run.inclusion_type == 'all' and not test_run.is_closed:
        logger.info(f"[TEST_RUN_DETAIL] No test cases found, trying to refresh results")
        print(f"[TEST_RUN_DETAIL] No test cases found, trying to refresh results")
        try:
            test_run.update_results_for_test_cases()
            test_cases = test_run.get_test_cases()
            if hasattr(test_cases, '__iter__') and not isinstance(test_cases, (list, tuple)):
                test_cases = list(test_cases)
            logger.info(f"[TEST_RUN_DETAIL] After refresh, found {len(test_cases)} test cases")
            print(f"[TEST_RUN_DETAIL] After refresh, found {len(test_cases)} test cases")
        except Exception as e:
            logger.error(f"[TEST_RUN_DETAIL] Error refreshing test cases for test run {test_run.id}: {str(e)}")
            print(f"[TEST_RUN_DETAIL] ERROR refreshing: {str(e)}")
    
    # Check results
    logger.info(f"[TEST_RUN_DETAIL] Test run has {len(test_run.results)} results")
    print(f"[TEST_RUN_DETAIL] Test run has {len(test_run.results)} results")
    
    results_map = {str(r.test_case.id): r for r in test_run.results}
    logger.info(f"[TEST_RUN_DETAIL] Results map has {len(results_map)} entries")
    print(f"[TEST_RUN_DETAIL] Results map has {len(results_map)} entries")
    
    # Get filter parameters
    assigned_filter = request.GET.get('assigned_to', '')
    status_filter = request.GET.get('status', '')
    section_filter = request.GET.get('section', '')
    
    # Filter test cases
    filtered_test_cases = []
    for test_case in test_cases:
        # Apply filters
        if assigned_filter:
            result = results_map.get(str(test_case.id))
            if not result or str(result.assigned_to_id) != assigned_filter:
                continue
        
        if status_filter:
            result = results_map.get(str(test_case.id))
            if not result or result.status != status_filter:
                continue
        
        if section_filter:
            if str(test_case.section.id) != section_filter:
                continue
        
        filtered_test_cases.append({
            'test_case': test_case,
            'result': results_map.get(str(test_case.id)),
        })
    
    # Get summary
    logger.info(f"[TEST_RUN_DETAIL] Getting results summary")
    print(f"[TEST_RUN_DETAIL] Getting results summary")
    summary = test_run.get_results_summary()
    logger.info(f"[TEST_RUN_DETAIL] Summary: {summary}")
    print(f"[TEST_RUN_DETAIL] Summary: total={summary['total']}, passed={summary['passed']}, failed={summary['failed']}, untested={summary['untested']}")
    
    # Get all sections for filter
    sections = Section.objects(project=project).order_by('name')
    
    # Get all users for filter
    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.filter(is_active=True).order_by('username')
    
    context = {
        'project': project,
        'test_run': test_run,
        'test_cases_data': filtered_test_cases,
        'summary': summary,
        'sections': sections,
        'users': users,
        'assigned_filter': assigned_filter,
        'status_filter': status_filter,
        'section_filter': section_filter,
    }
    return render(request, 'testcases/test_run_detail.html', context)


@login_required
def test_run_edit(request, project_id, test_run_id):
    """Edit a test run"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Validate test_run_id
    if not test_run_id:
        messages.error(request, 'Test run ID is required.')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    
    try:
        # Convert test_run_id to ObjectId
        test_run_id_str = str(test_run_id).strip()
        if not test_run_id_str:
            messages.error(request, 'Test run ID cannot be empty.')
            return redirect('testcases:test_run_list', project_id=str(project.id))
        
        try:
            test_run_object_id = ObjectId(test_run_id_str)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Invalid test run ID format: {test_run_id_str}, error: {str(e)}")
            messages.error(request, f'Invalid test run ID format: {test_run_id_str}')
            return redirect('testcases:test_run_list', project_id=str(project.id))
        
        test_run = TestRun.objects.get(id=test_run_object_id, project=project)
    except TestRun.DoesNotExist:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Test run not found: ID={test_run_id_str}, Project={project.id}")
        messages.error(request, f'Test run not found with ID: {test_run_id_str}')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading test run {test_run_id_str}: {str(e)}")
        messages.error(request, f'Error loading test run: {str(e)}')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        references = request.POST.get('references', '').strip()
        milestone_id = request.POST.get('milestone', '').strip()
        assigned_to_id = request.POST.get('assigned_to', '').strip()
        start_date_str = request.POST.get('start_date', '').strip()
        end_date_str = request.POST.get('end_date', '').strip()
        inclusion_type = request.POST.get('inclusion_type', 'all').strip()
        test_case_ids = []
        raw_sources = request.POST.getlist('test_case_ids[]')
        for source in raw_sources:
            for tc_id in (source or '').split(','):
                tc_id = tc_id.strip()
                if tc_id:
                    test_case_ids.append(tc_id)
        
        # Validate required fields
        if not name:
            messages.error(request, 'Test run name is required.')
            from django.contrib.auth import get_user_model
            User = get_user_model()
            users = User.objects.filter(is_active=True).order_by('username')
            context = {
                'project': project,
                'test_run': test_run,
                'users': users,
            }
            return render(request, 'testcases/test_run_edit.html', context)
        
        # Parse dates
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                pass
        
        # Parse assigned_to_id
        assigned_to = None
        if assigned_to_id:
            try:
                assigned_to = int(assigned_to_id)
            except ValueError:
                pass
        
        # Update test run
        test_run.name = name
        test_run.description = description
        test_run.references = references
        test_run.milestone_id = milestone_id
        test_run.assigned_to_id = assigned_to
        test_run.start_date = start_date
        test_run.end_date = end_date
        test_run.inclusion_type = inclusion_type
        test_run.test_case_ids = test_case_ids if inclusion_type == 'specific' else []
        test_run.updated_by_id = request.user.id
        test_run.save()
        
        # Update results for test cases
        test_run.update_results_for_test_cases()
        
        # Record change history
        record_change('test_run', str(test_run.id), 'updated', request.user.id, f'Updated test run: {test_run.name}')
        
        messages.success(request, 'Test run updated successfully.')
        return redirect('testcases:test_run_detail', project_id=str(project.id), test_run_id=str(test_run.id))
    
    # GET request - show form
    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.filter(is_active=True).order_by('username')
    
    context = {
        'project': project,
        'test_run': test_run,
        'users': users,
        'selected_case_ids_str': ','.join(test_run.test_case_ids or []),
        'selected_case_ids_json': json.dumps(test_run.test_case_ids or []),
    }
    return render(request, 'testcases/test_run_edit.html', context)


@login_required
def test_run_delete(request, project_id, test_run_id):
    """Delete a test run"""
    if not request.user.can_access_test_cases:
        messages.error(request, 'You do not have permission to access the Test Cases project.')
        return redirect('dashboard:home')
    
    project = get_project_or_redirect(project_id)
    if not project:
        messages.error(request, 'Project not found.')
        return redirect('testcases:project_dashboard')
    
    # Validate test_run_id
    if not test_run_id:
        messages.error(request, 'Test run ID is required.')
        return redirect('testcases:test_run_list', project_id=str(project.id))
    
    try:
        # Convert test_run_id to ObjectId
        test_run_id_str = str(test_run_id).strip()
        if not test_run_id_str:
            messages.error(request, 'Test run ID cannot be empty.')
            return redirect('testcases:test_run_list', project_id=str(project.id))
        
        try:
            test_run_object_id = ObjectId(test_run_id_str)
        except Exception:
            messages.error(request, f'Invalid test run ID format: {test_run_id_str}')
            return redirect('testcases:test_run_list', project_id=str(project.id))
        
        test_run = TestRun.objects.get(id=test_run_object_id, project=project)
        test_run_name = test_run.name
        test_run.delete()
        
        # Record change history
        record_change('test_run', test_run_id_str, 'deleted', request.user.id, f'Deleted test run: {test_run_name}')
        
        messages.success(request, f'Test run "{test_run_name}" deleted successfully!')
    except TestRun.DoesNotExist:
        messages.error(request, 'Test run not found.')
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting test run {test_run_id}: {str(e)}")
        messages.error(request, f'Error deleting test run: {str(e)}')
    
    return redirect('testcases:test_run_list', project_id=str(project.id))


@login_required
def test_run_select_cases(request, project_id):
    """AJAX endpoint to get test cases for selection modal"""
    if not request.user.can_access_test_cases:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    project = get_project_or_redirect(project_id)
    if not project:
        return JsonResponse({'error': 'Project not found'}, status=404)
    
    # Get all sections with their test cases, building hierarchical structure
    all_sections = Section.objects(project=project).order_by('name')
    
    # Build a map of sections
    section_map = {}
    root_sections = []
    
    # First pass: create section data structures
    for section in all_sections:
        test_cases = TestCase.objects(section=section, project=project, is_deleted=False).order_by('title')
        section_data = {
            'id': str(section.id),
            'name': section.name,
            'parent_id': str(section.parent.id) if section.parent else None,
            'test_cases': [
                {
                    'id': str(tc.id),
                    'title': tc.title,
                    'type': tc.type,
                    'priority': tc.priority,
                }
                for tc in test_cases
            ],
            'children': []
        }
        section_map[str(section.id)] = section_data
    
    # Second pass: build tree structure
    sections_data = []
    for section in all_sections:
        section_data = section_map[str(section.id)]
        if section.parent:
            parent_id = str(section.parent.id)
            if parent_id in section_map:
                section_map[parent_id]['children'].append(section_data)
        else:
            sections_data.append(section_data)
    
    # Sort sections and their children recursively
    def sort_sections(sections_list):
        sections_list.sort(key=lambda x: x['name'].lower())
        for section in sections_list:
            if section['children']:
                sort_sections(section['children'])
    
    sort_sections(sections_data)
    
    return JsonResponse({'sections': sections_data})


@login_required
def test_run_update_result(request, project_id, test_run_id):
    """Update test result for a test case in a test run"""
    if not request.user.can_access_test_cases:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    project = get_project_or_redirect(project_id)
    if not project:
        return JsonResponse({'error': 'Project not found'}, status=404)
    
    # Validate test_run_id
    if not test_run_id:
        return JsonResponse({'error': 'Test run ID is required'}, status=400)
    
    try:
        # Convert test_run_id to ObjectId
        test_run_id_str = str(test_run_id).strip()
        if not test_run_id_str:
            return JsonResponse({'error': 'Test run ID cannot be empty'}, status=400)
        
        try:
            test_run_object_id = ObjectId(test_run_id_str)
        except Exception:
            return JsonResponse({'error': f'Invalid test run ID format: {test_run_id_str}'}, status=400)
        
        test_run = TestRun.objects.get(id=test_run_object_id, project=project)
    except TestRun.DoesNotExist:
        return JsonResponse({'error': 'Test run not found'}, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading test run {test_run_id}: {str(e)}")
        return JsonResponse({'error': f'Error loading test run: {str(e)}'}, status=500)
    
    if request.method == 'POST':
        test_case_id = request.POST.get('test_case_id')
        status = request.POST.get('status')
        comment = request.POST.get('comment', '').strip()
        assigned_to_id = request.POST.get('assigned_to_id', '').strip()
        step_number = request.POST.get('step_number')  # Optional: for step-level results
        
        if not test_case_id or not status:
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        # Find or create result
        result = None
        for r in test_run.results:
            if str(r.test_case.id) == test_case_id:
                result = r
                break
        
        if not result:
            # Create new result
            test_case = resolve_test_case(project, test_case_id, include_deleted=True)
            if not test_case:
                return JsonResponse({'error': 'Test case not found'}, status=404)
            
            result = TestRunResult(
                test_case=test_case,
                status=status,
                comment=comment,
                assigned_to_id=int(assigned_to_id) if assigned_to_id else None,
                step_results={}
            )
            test_run.results.append(result)
        else:
            # Update existing result
            result.status = status
            result.comment = comment
            if assigned_to_id:
                result.assigned_to_id = int(assigned_to_id)
            result.updated_at = datetime.utcnow()
        
        # Store step-level result if step_number is provided
        if step_number:
            if not result.step_results:
                result.step_results = {}
            result.step_results[str(step_number)] = status
            
            # Determine overall test case status based on step results
            test_case = result.test_case
            if test_case and test_case.steps:
                total_steps = len(test_case.steps)
                step_results = result.step_results or {}
                
                # Count step statuses
                passed_steps = sum(1 for s in step_results.values() if s == 'passed')
                failed_steps = sum(1 for s in step_results.values() if s == 'failed')
                blocked_steps = sum(1 for s in step_results.values() if s == 'blocked')
                retest_steps = sum(1 for s in step_results.values() if s == 'retest')
                steps_with_results = len(step_results)  # Number of steps that have been evaluated
                
                # Determine overall status based on step results
                # Priority: failed > blocked > retest > passed (only if all steps passed)
                if failed_steps > 0:
                    result.status = 'failed'
                elif blocked_steps > 0:
                    result.status = 'blocked'
                elif retest_steps > 0:
                    result.status = 'retest'
                elif passed_steps == total_steps and steps_with_results == total_steps:
                    # Only mark as passed if ALL steps have results AND all are passed
                    result.status = 'passed'
                elif passed_steps > 0 and steps_with_results < total_steps:
                    # Some steps passed but not all steps have results yet - set to retest
                    result.status = 'retest'
                elif passed_steps > 0 and steps_with_results == total_steps:
                    # All steps have results but not all passed - should not happen if logic is correct
                    # But if it does, set to retest
                    result.status = 'retest'
                else:
                    # No steps passed or no results yet - keep as untested or current status
                    if result.status == 'untested' and steps_with_results > 0:
                        # Some steps evaluated but none passed - could be retest
                        result.status = 'retest'
                    # Otherwise keep current status
        else:
            # If no step_number, update overall status directly
            result.status = status
        
        test_run.save()
        
        # Get updated summary
        summary = test_run.get_results_summary()
        
        return JsonResponse({
            'success': True,
            'summary': summary,
        })
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def test_run_get_test_case_details(request, project_id, test_run_id, test_case_id):
    """Get test case details for display in test run detail view"""
    if not request.user.can_access_test_cases:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    project = get_project_or_redirect(project_id)
    if not project:
        return JsonResponse({'error': 'Project not found'}, status=404)
    
    try:
        test_run = TestRun.objects.get(id=ObjectId(test_run_id), project=project)
        test_case = resolve_test_case(project, test_case_id)
        
        if not test_case:
            return JsonResponse({'error': 'Test case not found'}, status=404)
        
        # Get result for this test case in this test run
        result = None
        for r in test_run.results:
            if str(r.test_case.id) == str(test_case.id):
                result = r
                break
        
        # Get steps
        steps = sorted(test_case.steps, key=lambda x: x.step_number) if test_case.steps else []
        
        # Prepare response
        response_data = {
            'id': str(test_case.id),
            'title': test_case.title,
            'type': test_case.type,
            'priority': test_case.priority,
            'description': test_case.description or '',
            'preconditions': test_case.preconditions or '',
            'section': {
                'id': str(test_case.section.id),
                'name': test_case.section.name,
            },
            'steps': [
                {
                    'step_number': step.step_number,
                    'description': step.description,
                    'expected_result': step.expected_result,
                    'status': result.step_results.get(str(step.step_number), 'untested') if result and result.step_results else 'untested',
                }
                for step in steps
            ],
            'result': {
                'status': result.status if result else 'untested',
                'comment': result.comment if result else '',
                'assigned_to_id': result.assigned_to_id if result and result.assigned_to_id else None,
                'assigned_to_username': result.assigned_to.username if result and result.assigned_to else None,
                'updated_at': result.updated_at.isoformat() if result and result.updated_at else None,
            } if result else {
                'status': 'untested',
                'comment': '',
                'assigned_to_id': None,
                'assigned_to_username': None,
                'updated_at': None,
            }
        }
        
        return JsonResponse(response_data)
        
    except TestRun.DoesNotExist:
        return JsonResponse({'error': 'Test run not found'}, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error getting test case details: {str(e)}")
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)

